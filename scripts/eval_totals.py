"""
eval_totals.py — Independently recompute PAIE and COMPTA totals from source files
and compare to what is written in the FT-P-2 workbook TOTAL rows.

Column positions are derived at runtime from the Feuil2 header row via col_utils —
if columns move, this script adapts automatically.

Exit code: 0 = all pass, 1 = mismatches found.
"""
import json, sys
import pandas as pd
from openpyxl import load_workbook
from col_utils import get_feuil2_col_map

TOLERANCE = 1  # FCFA

SESSION_FILE = ".audit-session.json"
with open(SESSION_FILE, encoding="utf-8") as f:
    session = json.load(f)

FT_PATH = session["files"]["feuille_travail"]
BG_PATH = session["files"]["balance_generale"]
LP_PATH = ".livre_paie_pivot.csv"
CP_PATH = ".charges_patronales_pivot.csv"

# Load workbook to resolve column map from header row
wb = load_workbook(FT_PATH, data_only=True)
ws = wb["Feuil2"]
col_map = get_feuil2_col_map(ws, header_row=3)

# ── Recompute expected totals ───────────────────────────────────────────────
lp = pd.read_csv(LP_PATH, dtype={"Matricule": str})
cp = pd.read_csv(CP_PATH, dtype={"Matricule": str})

def cp_sum(fragment):
    """Sum the first charges patronales column whose name contains `fragment`."""
    matched = [c for c in cp.columns if fragment.upper() in c.upper()]
    return round(cp[matched[0]].sum()) if matched else 0

expected_paie = {
    "SAL_BRUT":  round(lp["SAL BRUT"].sum()) if "SAL BRUT" in lp.columns else 0,
    "CNPS_P":    cp_sum("CNPS"),
    "CF_P":      cp_sum("CF/P"),
    "FNE":       cp_sum("FNE"),
    "AF":        cp_sum("AF"),
    "AT":        cp_sum("AT"),
}
expected_paie["CF_FNE"]    = expected_paie["CF_P"] + expected_paie["FNE"]
expected_paie["TOTAL_COL"] = sum(
    expected_paie[k] for k in ["SAL_BRUT", "CNPS_P", "CF_P", "FNE", "AF", "AT"]
)

# COMPTA from Balance Générale (MvtDebit col4 – MvtCredit col5)
df_bg = pd.read_excel(BG_PATH, dtype={"Compte": str})
df_bg.columns = [str(c).strip() for c in df_bg.columns]
compte_col = df_bg.columns[0]
mvtd_col   = df_bg.columns[4] if len(df_bg.columns) > 4 else None
mvtc_col   = df_bg.columns[5] if len(df_bg.columns) > 5 else None
df_bg[compte_col] = df_bg[compte_col].astype(str).str.strip()
df_bg["NetSolde"] = (
    pd.to_numeric(df_bg[mvtd_col], errors="coerce").fillna(0) -
    pd.to_numeric(df_bg[mvtc_col], errors="coerce").fillna(0)
)
account_map = {row[compte_col]: round(row["NetSolde"]) for _, row in df_bg.iterrows()}

# SYSCOHADA: accounts → logical column name
ACCT_TO_LOGICAL = {
    # 661x + 663x → Section A (SAL BRUT)
    **{acct: "SAL_BRUT" for acct in [
        "661110","661120","661130","661200","661210","661220",
        "661300","661380","661410","661800","663101","663102","663410"
    ]},
    "664120": "CNPS_P",   # Cotisations CNPS patronales
    "664110": "AF",        # Allocations Familiales
    "664130": "AT",        # Accident de Travail
    "664380": "CF_P",      # Crédit Foncier Patronal
    # FNE (664400) is structural zero — never booked in GL
}

expected_compta = {k: 0 for k in ["SAL_BRUT","CNPS_P","CF_P","FNE","CF_FNE","AF","AT","TOTAL_COL"]}
for acct, logical in ACCT_TO_LOGICAL.items():
    expected_compta[logical] += account_map.get(acct, 0)
expected_compta["CF_FNE"]    = expected_compta["CF_P"] + expected_compta["FNE"]
expected_compta["TOTAL_COL"] = sum(
    expected_compta[k] for k in ["SAL_BRUT", "CNPS_P", "CF_P", "FNE", "AF", "AT"]
)

# ── Read actual totals from workbook (by logical name → col index) ──────────
tot_paie_row   = session.get("tot_paie_row", 178)
tot_compta_row = session.get("tot_compta_row")

LOGICAL_NAMES = ["SAL_BRUT","CNPS_P","CF_P","FNE","CF_FNE","AF","AT","TOTAL_COL"]

def read_logical_row(ws, row_num, col_map, names):
    return {
        name: (ws.cell(row_num, col_map[name]).value or 0)
        for name in names if name in col_map
    }

actual_paie   = read_logical_row(ws, tot_paie_row, col_map, LOGICAL_NAMES)
actual_compta = read_logical_row(ws, tot_compta_row, col_map, LOGICAL_NAMES) if tot_compta_row else {}
wb.close()

# ── Compare ──────────────────────────────────────────────────────────────────
failures = []
print("=" * 70)
print("eval_totals — Verification Report (column-name-driven)")
print("=" * 70)

print(f"\nPAIE (row {tot_paie_row}):")
for name in LOGICAL_NAMES:
    if name not in col_map:
        continue
    exp = expected_paie.get(name, 0)
    act = actual_paie.get(name, 0)
    try:
        act_f = float(act)
    except (TypeError, ValueError):
        act_f = 0.0
    diff = abs(act_f - exp)
    status = "OK" if diff <= TOLERANCE else "FAIL"
    print(f"  [{status}] {name:15s}: expected={exp:>20,.0f}  actual={act_f:>20,.0f}  diff={diff:,.0f}")
    if diff > TOLERANCE:
        failures.append(
            f"PAIE row {tot_paie_row} {name}: expected {exp:,} got {act_f:,} diff {diff:,}"
        )

if tot_compta_row:
    print(f"\nCOMPTA (row {tot_compta_row}):")
    for name in LOGICAL_NAMES:
        if name not in col_map:
            continue
        exp = expected_compta.get(name, 0)
        act = actual_compta.get(name, 0)
        try:
            act_f = float(act)
        except (TypeError, ValueError):
            act_f = 0.0
        diff = abs(act_f - exp)
        status = "OK" if diff <= TOLERANCE else "FAIL"
        print(f"  [{status}] {name:15s}: expected={exp:>20,.0f}  actual={act_f:>20,.0f}  diff={diff:,.0f}")
        if diff > TOLERANCE:
            failures.append(
                f"COMPTA row {tot_compta_row} {name}: expected {exp:,} got {act_f:,} diff {diff:,}"
            )

print("\n" + "=" * 70)
if failures:
    print(f"FAIL -- {len(failures)} mismatch(es) found")
    for f in failures:
        print(f"   * {f}")
    sys.exit(1)
else:
    print("PASS -- All column totals match source data")
    sys.exit(0)
