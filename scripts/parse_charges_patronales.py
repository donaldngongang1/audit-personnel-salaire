"""
parse_charges_patronales.py — Parse Charges Patronales CSV → TCD/pivot format.
Output: Extract Charges Patronal sheet — one row per employee, 5 patronal charge codes.
Final rows: 'Total' row (= column sums) + 'TOTAL' row (= 2× for patronal+salarial).

Filter: Matricule must match ^\\d{3,} (exclude "Total"/"TOTAL" summary rows).
"""
import io, json, re, sys
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

SESSION_FILE = ".audit-session.json"
with open(SESSION_FILE, encoding="utf-8") as f:
    session = json.load(f)

CSV_PATH = session["files"]["charges_patronales"]
FT_PATH  = session["files"]["feuille_travail"]

print(f"Parsing Charges Patronales: {CSV_PATH}")

def parse_csv_value(raw):
    s = raw.strip()
    s = re.sub(r'^=""?', '', s)
    s = re.sub(r'""?$', '', s)
    return s

def parse_amount(raw):
    s = parse_csv_value(raw)
    s = s.replace('","', '.').replace(',', '.').replace(' ', '')
    try:
        return float(s)
    except ValueError:
        return 0.0

MATRICULE_RE = re.compile(r'^\d{3,}')

CODE_MAP = {
    "4100": "Crédit Foncier Patronal",
    "4400": "Fond National de l'emploi (FNE)",
    "4500": "Pension Vieillesse (CNPS)",
    "4800": "Allocation Familiale",
    "4900": "Accident de Travail",
}

rows = []
excluded_total = 0
excluded_unknown = 0

with open(CSV_PATH, encoding="latin-1") as f:
    for line in f:
        line = line.rstrip('\n\r')
        if not line.strip():
            continue
        parts = line.split(';')
        if len(parts) < 6:
            continue
        matricule = parse_csv_value(parts[0])
        if not MATRICULE_RE.match(matricule):
            excluded_total += 1
            continue
        nom       = parse_csv_value(parts[1])
        prenom    = parse_csv_value(parts[2]) if len(parts) > 2 else ""
        type_code = parse_csv_value(parts[4]) if len(parts) > 4 else ""
        if type_code not in CODE_MAP:
            excluded_unknown += 1
            continue
        amount = parse_amount(parts[5]) if len(parts) > 5 else 0.0
        rows.append({
            "Matricule": matricule,
            "Nom": nom,
            "Prenom": prenom,
            "TypeCode": type_code,
            "Montant": amount,
        })

print(f"  Excluded: {excluded_total} Total rows | {excluded_unknown} unknown codes | {len(rows)} rows kept")

df = pd.DataFrame(rows)
if df.empty:
    print("⚠️ 0 rows found after filtering — verify CSV format")
    sys.exit(1)

df["ColName"] = df["TypeCode"].map(CODE_MAP)

# Pivot: one row per employee, one column per charge type
pivot = df.pivot_table(
    index=["Matricule", "Nom", "Prenom"],
    columns="ColName",
    values="Montant",
    aggfunc="sum",
    fill_value=0,
).reset_index()
pivot.columns.name = None
pivot = pivot.sort_values("Matricule").reset_index(drop=True)

# Ensure all 5 charge columns exist in correct order
ORDERED_CHARGES = [
    "Crédit Foncier Patronal",
    "Fond National de l'emploi (FNE)",
    "Pension Vieillesse (CNPS)",
    "Allocation Familiale",
    "Accident de Travail",
]
for col in ORDERED_CHARGES:
    if col not in pivot.columns:
        pivot[col] = 0.0

final_cols = ["Matricule", "Nom", "Prenom"] + ORDERED_CHARGES
pivot = pivot[final_cols]

# Compute column sums (for "Total" and "TOTAL" rows)
col_sums = {col: pivot[col].sum() for col in ORDERED_CHARGES}
print(f"  Pivot: {len(pivot)} employees")
for col, s in col_sums.items():
    print(f"    {col}: {s:,.0f} FCFA")

# ── Write Extract Charges Patronal sheet ──────────────────────────────────────
wb = load_workbook(FT_PATH)
if "Extract Charges Patronal" in wb.sheetnames:
    del wb["Extract Charges Patronal"]
ws = wb.create_sheet("Extract Charges Patronal")

BLUE_FILL = PatternFill("solid", start_color="1F4E79")
WHT_FONT  = Font(name="Arial", bold=True, color="FFFFFF", size=10)
DAT_FONT  = Font(name="Arial", size=10)
TOT_FONT  = Font(name="Arial", bold=True, size=10, color="1F4E79")
TOT_FONT2 = Font(name="Arial", bold=True, size=10, color="FFFFFF")
TOT_FILL  = PatternFill("solid", start_color="D6E4F0")
TOT_FILL2 = PatternFill("solid", start_color="1F4E79")
ALT_FILL  = PatternFill("solid", start_color="F5F8FC")
RGT = Alignment(horizontal="right", vertical="center")
LFT = Alignment(horizontal="left",  vertical="center")
CTR = Alignment(horizontal="center", vertical="center")
NUM_FMT = "#,##0;(#,##0);\"-\""

def thn(): return Side(style="thin", color="D9D9D9")
def med(): return Side(style="medium", color="1F4E79")
BDR   = Border(bottom=thn(), right=thn())
BDR_T = Border(top=med(), bottom=med(), right=thn())

# Headers (matching TCD format)
headers = ["Etiquette de lignes", "NOM", "PRENOM"] + ORDERED_CHARGES
for c, h in enumerate(headers, 1):
    cell = ws.cell(1, c, value=h)
    cell.font = WHT_FONT; cell.fill = BLUE_FILL
    cell.alignment = LFT; cell.border = BDR

numeric_cols = set(range(4, 4 + len(ORDERED_CHARGES)))

# Employee rows
for i, row in enumerate(pivot.itertuples(index=False), 2):
    alt = ALT_FILL if (i % 2 == 0) else None
    vals = [row.Matricule, row.Nom, row.Prenom] + [getattr(row, c.replace(" ", "_").replace("(", "").replace(")", "").replace("'", "_")) if hasattr(row, c.replace(" ", "_").replace("(", "").replace(")", "").replace("'", "_")) else 0 for c in ORDERED_CHARGES]
    # Access pivot values by position
    row_vals = list(row)
    for c_idx, val in enumerate(row_vals, 1):
        cell = ws.cell(i, c_idx, value=val)
        cell.font = DAT_FONT; cell.border = BDR
        if alt: cell.fill = alt
        cell.alignment = RGT if c_idx in numeric_cols else LFT
        if c_idx in numeric_cols: cell.number_format = NUM_FMT

# "Total" row (= column sums, with "=" in NOM/PRENOM)
total_row1 = len(pivot) + 2
ws.cell(total_row1, 1, value="Total").font = TOT_FONT
ws.cell(total_row1, 1).fill = TOT_FILL
ws.cell(total_row1, 1).alignment = LFT
ws.cell(total_row1, 1).border = BDR_T
for c_label, label_val in [(2, "="), (3, "=")]:
    cell = ws.cell(total_row1, c_label, value=label_val)
    cell.font = TOT_FONT; cell.fill = TOT_FILL
    cell.alignment = CTR; cell.border = BDR_T
for c_offset, col_name in enumerate(ORDERED_CHARGES, 4):
    cell = ws.cell(total_row1, c_offset, value=col_sums[col_name])
    cell.font = TOT_FONT; cell.fill = TOT_FILL
    cell.number_format = NUM_FMT; cell.alignment = RGT; cell.border = BDR_T

# "TOTAL" row (= 2× column sums — patronal + salarial)
total_row2 = len(pivot) + 3
ws.cell(total_row2, 1, value="TOTAL").font = TOT_FONT2
ws.cell(total_row2, 1).fill = TOT_FILL2
ws.cell(total_row2, 1).alignment = LFT
ws.cell(total_row2, 1).border = BDR_T
for c_label in [2, 3]:
    cell = ws.cell(total_row2, c_label, value=None)
    cell.font = TOT_FONT2; cell.fill = TOT_FILL2; cell.border = BDR_T
for c_offset, col_name in enumerate(ORDERED_CHARGES, 4):
    cell = ws.cell(total_row2, c_offset, value=col_sums[col_name] * 2)
    cell.font = TOT_FONT2; cell.fill = TOT_FILL2
    cell.number_format = NUM_FMT; cell.alignment = RGT; cell.border = BDR_T

ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"
ws.freeze_panes = "A2"
ws.column_dimensions["A"].width = 22
ws.column_dimensions["B"].width = 25
ws.column_dimensions["C"].width = 20
for i in range(4, len(headers) + 1):
    ws.column_dimensions[get_column_letter(i)].width = 24

wb.save(FT_PATH)
print(f"✅ Extract Charges Patronal: {len(pivot)} employees + Total + TOTAL rows → '{FT_PATH}'")

# Persist pivot for reconciliation (save with safe column names)
pivot_save = pivot.copy()
pivot_save.columns = [c.replace("(", "").replace(")", "").replace("'", "").replace(" ", "_") for c in pivot_save.columns]
pivot_save.to_csv(".charges_patronales_pivot.csv", index=False, encoding="utf-8")

session.setdefault("extract_counts", {})["charges_patronales"] = len(pivot)
session.setdefault("steps_completed", [])
if "extract_charges_patronal" not in session["steps_completed"]:
    session["steps_completed"].append("extract_charges_patronal")
with open(SESSION_FILE, "w", encoding="utf-8") as f:
    json.dump(session, f, indent=2, ensure_ascii=False)
