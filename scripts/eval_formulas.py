"""
eval_formulas.py — Verify formula integrity in Feuil2 (column-name-driven).

Column positions are derived at runtime from the Feuil2 header row via col_utils.
If columns move, patterns adapt automatically — no hardcoded letters.

Checks:
  1. Columns flagged as blank (SALAIRE BASE, ANCIENNETE, H.SUP, AUTRE GAIN) must be
     empty on all data rows.
  2. CF_FNE column (V) formula must be =<CF_P_col>{r}+<FNE_col>{r}.
  3. TOTAL_COL column (Y) formula must be the sum of SAL_BRUT+CNPS_P+CF_FNE+AF+AT
     columns (uses CF_FNE, not CF_P+FNE separately — avoids double-count).
  4. No formula anywhere in a data row may reference a blank column.

Exit code: 0 = pass, 1 = violations found.
"""
import json, sys, re
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter as gcl
from col_utils import get_feuil2_col_map, cols_for_row_total

SESSION_FILE = ".audit-session.json"
with open(SESSION_FILE, encoding="utf-8") as f:
    session = json.load(f)

FT_PATH      = session["files"]["feuille_travail"]
data_start   = session.get("data_start", 4)
tot_paie_row = session.get("tot_paie_row", 178)

# Load WITHOUT data_only so we can inspect formula strings
wb = load_workbook(FT_PATH)
ws = wb["Feuil2"]

col_map = get_feuil2_col_map(ws, header_row=3)

# ── Derive column letters from the map ─────────────────────────────────────
cf_p_letter  = gcl(col_map["CF_P"])
fne_letter   = gcl(col_map["FNE"])
v_col_idx    = col_map["CF_FNE"]
v_letter     = gcl(v_col_idx)
y_col_idx    = col_map["TOTAL_COL"]
y_letter     = gcl(y_col_idx)

# Letters that Y should sum (SAL_BRUT + CNPS_P + CF_FNE + AF + AT)
y_term_letters = [gcl(c) for c in cols_for_row_total(col_map)]
y_expected_str = "+".join(y_term_letters)

blank_col_indices = set(col_map.get("BLANK_COLS", []))
blank_col_letters = {gcl(c) for c in blank_col_indices}

# ── Build dynamic regex patterns ────────────────────────────────────────────
# V: =<CF_P>{r}+<FNE>{r}  — exactly two terms, any row number
V_PATTERN = re.compile(
    rf"={re.escape(cf_p_letter)}\d+\+{re.escape(fne_letter)}\d+",
    re.IGNORECASE,
)

# Y: =<term1>{r}+<term2>{r}+... — exactly the required terms in order
_y_pat = r"\+".join(rf"{re.escape(l)}\d+" for l in y_term_letters)
Y_PATTERN = re.compile(rf"={_y_pat}", re.IGNORECASE)

# Bad refs: any formula referencing a blank column
if blank_col_letters:
    _bad = "|".join(
        rf"[=+\-\*/]{re.escape(l)}\d+" for l in sorted(blank_col_letters)
    )
    BAD_REFS = re.compile(_bad)
else:
    BAD_REFS = None

# ── Print header ────────────────────────────────────────────────────────────
print("=" * 70)
print("eval_formulas -- Formula Integrity Check (column-name-driven)")
print(f"  Checking rows {data_start}-{tot_paie_row - 1}")
print(f"  V column ({v_letter:2s}): formula must be ={cf_p_letter}{{r}}+{fne_letter}{{r}}")
print(f"  Y column ({y_letter:2s}): formula must be ={y_expected_str} (using row numbers)")
if blank_col_letters:
    print(f"  Blank cols: {sorted(blank_col_letters)} (must stay empty)")
print("=" * 70)

blank_violations   = []
v_violations       = []
y_violations       = []
bad_ref_violations = []

for r in range(data_start, tot_paie_row):
    # 1. Blank column check
    for col in blank_col_indices:
        val = ws.cell(r, col).value
        if val not in (None, 0, ""):
            blank_violations.append(
                f"Row {r} Col {gcl(col)}: expected blank, got '{val}'"
            )

    # 2. V (CF_FNE) formula check
    v_val = str(ws.cell(r, v_col_idx).value or "")
    if not V_PATTERN.search(v_val):
        v_violations.append(
            f"Row {r} {v_letter}: '{v_val}'"
            f" (expected ={cf_p_letter}{r}+{fne_letter}{r})"
        )

    # 3. Y (TOTAL) formula check
    y_val = str(ws.cell(r, y_col_idx).value or "")
    if not Y_PATTERN.search(y_val):
        y_violations.append(
            f"Row {r} {y_letter}: '{y_val}'"
            f" (expected ={'+'.join(f'{l}{r}' for l in y_term_letters)})"
        )

    # 4. Stale blank-column references
    if BAD_REFS:
        for col in range(1, ws.max_column + 1):
            cell_val = str(ws.cell(r, col).value or "")
            if cell_val.startswith("=") and BAD_REFS.search(cell_val):
                bad_ref_violations.append(
                    f"Row {r} Col {gcl(col)}: '{cell_val}' references a blank column"
                )

# ── Report ──────────────────────────────────────────────────────────────────
if blank_violations:
    print(f"\nFAIL -- Blank column violations: {len(blank_violations)}")
    for v in blank_violations[:10]:
        print(f"   * {v}")
else:
    print(f"PASS -- Blank columns: all clear on {tot_paie_row - data_start} data rows")

if v_violations:
    print(f"\nFAIL -- {v_letter} (CF_FNE) formula issues: {len(v_violations)}")
    for v in v_violations[:10]:
        print(f"   * {v}")
else:
    print(f"PASS -- {v_letter} column: all formulas = {cf_p_letter}+{fne_letter}")

if y_violations:
    print(f"\nFAIL -- {y_letter} (TOTAL) formula issues: {len(y_violations)}")
    for v in y_violations[:10]:
        print(f"   * {v}")
else:
    print(f"PASS -- {y_letter} column: all formulas = {y_expected_str}")

if bad_ref_violations:
    print(f"\nFAIL -- Stale blank-column references: {len(bad_ref_violations)}")
    for v in bad_ref_violations[:10]:
        print(f"   * {v}")
else:
    print("PASS -- No stale blank-column formula references found")

wb.close()
all_violations = blank_violations + v_violations + y_violations + bad_ref_violations

print("\n" + "=" * 70)
if all_violations:
    print(f"FAIL -- {len(all_violations)} total formula violation(s)")
    sys.exit(1)
else:
    print("PASS -- All formula integrity checks passed")
    sys.exit(0)
