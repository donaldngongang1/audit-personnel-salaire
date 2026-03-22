"""
parse_grand_livre.py — Parse Grand Livre Général (.xls) and write Extract GL sheet.
Filter: CodeJournal == 'CAM' AND account starts with '66'.
"""
import json, os, sys
import pandas as pd
import xlrd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

SESSION_FILE = ".audit-session.json"
with open(SESSION_FILE, encoding="utf-8") as f:
    session = json.load(f)

GL_PATH = session["files"]["grand_livre"]
FT_PATH = session["files"]["feuille_travail"]

print(f"Parsing Grand Livre Général: {GL_PATH}")

# Read .xls file (requires xlrd)
xls = xlrd.open_workbook(GL_PATH)
# Use first sheet or 'Sage' if present
sheet_names = xls.sheet_names()
sheet_name = "Sage" if "Sage" in sheet_names else sheet_names[0]
print(f"  Sheet: {sheet_name} | Total sheets: {sheet_names}")

df = pd.read_excel(GL_PATH, sheet_name=sheet_name, engine="xlrd", dtype=str)
df.columns = [str(c).strip() for c in df.columns]
print(f"  Columns: {list(df.columns)}")
print(f"  Total rows: {len(df)}")

# Detect key columns
compte_col = next((c for c in df.columns if c.lower() in ["compte", "n°compte", "n° compte", "account"]), df.columns[0])
journal_col = next((c for c in df.columns if "journal" in c.lower() or "codejournal" in c.lower().replace(" ","")), None)
if journal_col is None and len(df.columns) >= 3:
    journal_col = df.columns[2]

debit_col  = next((c for c in df.columns if "debit" in c.lower()), None)
credit_col = next((c for c in df.columns if "credit" in c.lower()), None)
solde_col  = next((c for c in df.columns if "solde" in c.lower()), None)

print(f"  Compte: {compte_col}, Journal: {journal_col}, Debit: {debit_col}, Credit: {credit_col}")

# Apply filters
mask_journal = df[journal_col].astype(str).str.strip().str.upper() == "CAM"
mask_account = df[compte_col].astype(str).str.strip().str.startswith("66")
df_filtered = df[mask_journal & mask_account].copy()

print(f"  Filtered (CAM + 66xxx): {len(df_filtered)} rows")

# Convert numeric cols
for col in [debit_col, credit_col, solde_col]:
    if col:
        df_filtered[col] = pd.to_numeric(df_filtered[col], errors="coerce").fillna(0)

# ── Write Extract GL sheet ──────────────────────────────────────────────────────
wb = load_workbook(FT_PATH)
if "Extract GL" in wb.sheetnames:
    del wb["Extract GL"]
ws = wb.create_sheet("Extract GL")

BLUE_FILL = PatternFill("solid", start_color="1F4E79")
WHITE_FONT = Font(name="Arial", bold=True, color="FFFFFF", size=10)
DATA_FONT  = Font(name="Arial", size=10)
ALT_FILL   = PatternFill("solid", start_color="F5F8FC")
RGT = Alignment(horizontal="right", vertical="center")
LFT = Alignment(horizontal="left",  vertical="center")
NUM_FMT = "#,##0;(#,##0);\"-\""

def thin(): return Side(style="thin", color="D9D9D9")
BDR = Border(bottom=thin(), right=thin())

# Write all columns from the filtered dataframe
headers = list(df_filtered.columns)
for col_idx, header in enumerate(headers, 1):
    c = ws.cell(row=1, column=col_idx, value=header)
    c.font = WHITE_FONT; c.fill = BLUE_FILL; c.alignment = LFT; c.border = BDR

numeric_cols = {debit_col, credit_col, solde_col}
for r_idx, row in enumerate(df_filtered.itertuples(index=False), 2):
    alt = ALT_FILL if (r_idx % 2 == 0) else None
    for c_idx, (val, col_name) in enumerate(zip(row, headers), 1):
        cell = ws.cell(row=r_idx, column=c_idx, value=val)
        cell.font = DATA_FONT; cell.border = BDR
        if alt: cell.fill = alt
        if col_name in numeric_cols:
            cell.number_format = NUM_FMT; cell.alignment = RGT
        else:
            cell.alignment = LFT

ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"
ws.freeze_panes = "A2"

# Column widths
for i, col_name in enumerate(headers, 1):
    cl = get_column_letter(i)
    ws.column_dimensions[cl].width = max(12, min(35, len(col_name) + 4))

wb.save(FT_PATH)
print(f"✅ Extract GL: {len(df_filtered)} rows written to '{FT_PATH}'")

session.setdefault("extract_counts", {})["gl"] = len(df_filtered)
session.setdefault("steps_completed", [])
if "extract_gl" not in session["steps_completed"]:
    session["steps_completed"].append("extract_gl")
with open(SESSION_FILE, "w", encoding="utf-8") as f:
    json.dump(session, f, indent=2, ensure_ascii=False)
