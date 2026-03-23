"""
eval_ecart.py — Verify that ECART row = TOTAL COMPTA - TOTAL PAIE for each column.

Column positions are derived at runtime from the Feuil2 header row via col_utils —
if columns move, this script adapts automatically.

Exit code: 0 = pass, 1 = failures.
"""
import json, sys
from openpyxl import load_workbook
from col_utils import get_feuil2_col_map

TOLERANCE = 1

SESSION_FILE = ".audit-session.json"
with open(SESSION_FILE, encoding="utf-8") as f:
    session = json.load(f)

FT_PATH        = session["files"]["feuille_travail"]
tot_paie_row   = session.get("tot_paie_row", 178)
tot_compta_row = session.get("tot_compta_row")
ecart_row      = session.get("ecart_row")

wb = load_workbook(FT_PATH, data_only=True)
ws = wb["Feuil2"]

col_map = get_feuil2_col_map(ws, header_row=3)

# Scan for TOTAL COMPTA / ECART rows if not in session
if not tot_compta_row:
    print("Warning: TOTAL COMPTA row not in session -- scanning Feuil2...")
    for r in range(1, ws.max_row + 1):
        v = str(ws.cell(r, 1).value or ws.cell(r, 2).value or "")
        if "TOTAL COMPTABILITE" in v.upper():
            tot_compta_row = r
            break

if not ecart_row:
    for r in range(1, ws.max_row + 1):
        v = str(ws.cell(r, 1).value or ws.cell(r, 2).value or "")
        if "ECART" in v.upper() and "TOTAL" in v.upper():
            ecart_row = r
            break

LOGICAL_NAMES = ["SAL_BRUT","CNPS_P","CF_P","FNE","CF_FNE","AF","AT","TOTAL_COL"]
failures = []

print("=" * 70)
print("eval_ecart -- ECART Row Verification (column-name-driven)")
print(f"  TOTAL PAIE row:   {tot_paie_row}")
print(f"  TOTAL COMPTA row: {tot_compta_row}")
print(f"  ECART row:        {ecart_row}")
print("=" * 70)

if not (tot_compta_row and ecart_row):
    print("FAIL -- Could not locate TOTAL COMPTA or ECART row in Feuil2")
    wb.close()
    sys.exit(1)

for name in LOGICAL_NAMES:
    if name not in col_map:
        continue
    col_idx = col_map[name]
    paie   = ws.cell(tot_paie_row,   col_idx).value or 0
    compta = ws.cell(tot_compta_row, col_idx).value or 0
    ecart  = ws.cell(ecart_row,      col_idx).value or 0
    try:
        paie_f   = float(paie)
        compta_f = float(compta)
        ecart_f  = float(ecart)
    except (TypeError, ValueError):
        paie_f = compta_f = ecart_f = 0.0

    expected_ecart = compta_f - paie_f
    diff = abs(ecart_f - expected_ecart)
    status = "OK" if diff <= TOLERANCE else "FAIL"
    print(
        f"  [{status}] {name:15s}: COMPTA={compta_f:>20,.0f}  PAIE={paie_f:>20,.0f}  "
        f"Expected ECART={expected_ecart:>15,.0f}  Actual={ecart_f:>15,.0f}  diff={diff:,.0f}"
    )
    if diff > TOLERANCE:
        failures.append(f"{name}: expected ecart {expected_ecart:,} got {ecart_f:,}")

wb.close()

print("\n" + "=" * 70)
if failures:
    print(f"FAIL -- {len(failures)} ecart mismatch(es)")
    for f in failures:
        print(f"   * {f}")
    sys.exit(1)
else:
    print("PASS -- All ECART cells = COMPTA - PAIE")
    sys.exit(0)
