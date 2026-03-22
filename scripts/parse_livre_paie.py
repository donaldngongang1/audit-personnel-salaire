"""
parse_livre_paie.py — Parse Livre de Paie CSV → TCD/pivot format.
Output: Extract LivrePaie sheet — one row per employee, rubrique BRUT only.
Final rows: 'TOTAL' row with grand total.

Filter: Matricule must match ^\\d{3,} (exclude "Total"/"TOTAL" summary rows).
Code match: exact TypeCode == "BRUT" only.
"""
import io, json, re, sys
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

SESSION_FILE = ".audit-session.json"
with open(SESSION_FILE, encoding="utf-8") as f:
    session = json.load(f)

CSV_PATH = session["files"]["livre_paie"]
FT_PATH  = session["files"]["feuille_travail"]

print(f"Parsing Livre de Paie: {CSV_PATH}")

# ── Custom CSV parser ─────────────────────────────────────────────────────────
def parse_csv_value(raw):
    s = raw.strip()
    s = re.sub(r'^=""?', '', s)
    s = re.sub(r'""?$', '', s)
    return s

def parse_amount(raw):
    s = parse_csv_value(raw)
    s = s.replace('","', '.').replace(',', '.').replace(' ', '')
    try:
        return float(s)
    except ValueError:
        return 0.0

MATRICULE_RE = re.compile(r'^\d{3,}')

rows = []
raw_counts = {"total_lines": 0, "excluded_total": 0, "excluded_non_brut": 0, "included": 0}

with open(CSV_PATH, encoding="latin-1") as f:
    for line in f:
        line = line.rstrip('\n\r')
        if not line.strip():
            continue
        parts = line.split(';')
        if len(parts) < 6:
            continue
        raw_counts["total_lines"] += 1
        matricule = parse_csv_value(parts[0])
        if not MATRICULE_RE.match(matricule):
            raw_counts["excluded_total"] += 1
            continue
        nom        = parse_csv_value(parts[1])
        prenom     = parse_csv_value(parts[2]) if len(parts) > 2 else ""
        type_code  = parse_csv_value(parts[4]) if len(parts) > 4 else ""
        if type_code != "BRUT":
            raw_counts["excluded_non_brut"] += 1
            continue
        amount = parse_amount(parts[5]) if len(parts) > 5 else 0.0
        rows.append({"Matricule": matricule, "Nom": nom, "Prenom": prenom, "SAL_BRUT": amount})
        raw_counts["included"] += 1

print(f"  Lines: {raw_counts['total_lines']} total | {raw_counts['excluded_total']} excluded (Total rows) | "
      f"{raw_counts['excluded_non_brut']} non-BRUT | {raw_counts['included']} BRUT rows")

df = pd.DataFrame(rows)
if df.empty:
    print("⚠️ 0 BRUT rows found — verify Matricule format and TypeCode values")
    sys.exit(1)

# Pivot: one row per employee
pivot = df.groupby(["Matricule", "Nom", "Prenom"], as_index=False)["SAL_BRUT"].sum()
pivot = pivot.sort_values("Matricule").reset_index(drop=True)
grand_total = pivot["SAL_BRUT"].sum()

print(f"  Pivot: {len(pivot)} employees | Total SAL BRUT: {grand_total:,.0f} FCFA")

# ── Write Extract LivrePaie sheet ─────────────────────────────────────────────
wb = load_workbook(FT_PATH)
if "Extract LivrePaie" in wb.sheetnames:
    del wb["Extract LivrePaie"]
ws = wb.create_sheet("Extract LivrePaie")

BLUE_FILL  = PatternFill("solid", start_color="1F4E79")
WHT_FONT   = Font(name="Arial", bold=True, color="FFFFFF", size=10)
DAT_FONT   = Font(name="Arial", size=10)
TOT_FONT   = Font(name="Arial", bold=True, size=10, color="1F4E79")
TOT_FILL   = PatternFill("solid", start_color="D6E4F0")
ALT_FILL   = PatternFill("solid", start_color="F5F8FC")
RGT = Alignment(horizontal="right", vertical="center")
LFT = Alignment(horizontal="left",  vertical="center")
NUM_FMT = "#,##0;(#,##0);\"-\""

def thn(): return Side(style="thin", color="D9D9D9")
def med(): return Side(style="medium", color="1F4E79")
BDR = Border(bottom=thn(), right=thn())
BDR_T = Border(top=med(), bottom=med(), right=thn())

# Headers (matching TCD format expected by template)
headers = ["Etiquette de lignes", "NOM", "PRENOM", "Salaire BRUT"]
for c, h in enumerate(headers, 1):
    cell = ws.cell(1, c, value=h)
    cell.font = WHT_FONT; cell.fill = BLUE_FILL
    cell.alignment = LFT; cell.border = BDR

# Employee rows
for i, row in enumerate(pivot.itertuples(index=False), 2):
    alt = ALT_FILL if (i % 2 == 0) else None
    for c, val in enumerate([row.Matricule, row.Nom, row.Prenom, row.SAL_BRUT], 1):
        cell = ws.cell(i, c, value=val)
        cell.font = DAT_FONT; cell.border = BDR
        if alt: cell.fill = alt
        cell.alignment = RGT if c == 4 else LFT
        if c == 4: cell.number_format = NUM_FMT

# TOTAL row
total_row = len(pivot) + 2
for c, val in enumerate(["TOTAL", None, None, grand_total], 1):
    cell = ws.cell(total_row, c, value=val)
    cell.font = TOT_FONT; cell.fill = TOT_FILL; cell.border = BDR_T
    cell.alignment = RGT if c == 4 else LFT
    if c == 4: cell.number_format = NUM_FMT

ws.auto_filter.ref = f"A1:D1"
ws.freeze_panes = "A2"
ws.column_dimensions["A"].width = 22
ws.column_dimensions["B"].width = 25
ws.column_dimensions["C"].width = 20
ws.column_dimensions["D"].width = 20

wb.save(FT_PATH)
print(f"✅ Extract LivrePaie: {len(pivot)} employees + TOTAL row → '{FT_PATH}'")

# Persist pivot for reconciliation
pivot.to_csv(".livre_paie_pivot.csv", index=False, encoding="utf-8")

session.setdefault("extract_counts", {})["livre_paie"] = len(pivot)
session.setdefault("steps_completed", [])
if "extract_livre_paie" not in session["steps_completed"]:
    session["steps_completed"].append("extract_livre_paie")
with open(SESSION_FILE, "w", encoding="utf-8") as f:
    json.dump(session, f, indent=2, ensure_ascii=False)
