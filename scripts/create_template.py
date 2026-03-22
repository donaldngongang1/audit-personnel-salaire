"""
create_template.py — Generate blank FT-P-2-template.xlsx with Feuil1, Feuil2, Feuil3.
Run once to create the assets/FT-P-2-template.xlsx file.
"""
import io, os, sys
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter as gcl

wb = Workbook()

# Remove default sheet
if "Sheet" in wb.sheetnames:
    del wb["Sheet"]

# ─────────────────────────────────────────────────────────────────
# Styling helpers
# ─────────────────────────────────────────────────────────────────
def med(c="1F4E79"):  return Side(style="medium",  color=c)
def thn(c="D9D9D9"):  return Side(style="thin",    color=c)

BDR_DATA  = Border(bottom=thn(), right=thn())
BDR_HEAVY = Border(top=med(), bottom=med(), left=med(), right=med())
BDR_TOT   = Border(top=med(), bottom=med(), right=thn())

RGT = Alignment(horizontal="right",  vertical="center", wrap_text=False)
CTR = Alignment(horizontal="center", vertical="center", wrap_text=True)
LFT = Alignment(horizontal="left",   vertical="center", wrap_text=True)

F_TITLE  = Font(name="Arial", bold=True, size=14, color="1F4E79")
F_HEADER = Font(name="Arial", bold=True, size=10, color="FFFFFF")
F_SUBHDR = Font(name="Arial", bold=True, size=10, color="1F4E79")
F_DATA   = Font(name="Arial", size=10)
F_TOT_P  = Font(name="Arial", bold=True, size=10, color="1F4E79")
F_TOT_C  = Font(name="Arial", bold=True, size=10, color="1F4E79")
F_ECART  = Font(name="Arial", bold=True, size=10, color="FFFFFF")

FILL_BLUE   = PatternFill("solid", start_color="1F4E79")
FILL_ALT    = PatternFill("solid", start_color="F5F8FC")
FILL_TOT_P  = PatternFill("solid", start_color="D6E4F0")
FILL_TOT_C  = PatternFill("solid", start_color="C6EFCE")
FILL_ECART  = PatternFill("solid", start_color="843C0C")
FILL_GREY   = PatternFill("solid", start_color="D9D9D9")
FILL_YELLOW = PatternFill("solid", start_color="FFFF00")
FILL_NONE   = PatternFill(fill_type=None)

NUM_FMT = "#,##0;(#,##0);\"-\""

def hcell(ws, r, c, val, font=None, fill=None, align=None, bdr=None, num_fmt=None):
    cell = ws.cell(r, c, value=val)
    if font:    cell.font       = font
    if fill:    cell.fill       = fill
    if align:   cell.alignment  = align
    if bdr:     cell.border     = bdr
    if num_fmt: cell.number_format = num_fmt
    return cell

# ─────────────────────────────────────────────────────────────────
# FEUIL1 — Executive Summary Sheet
# ─────────────────────────────────────────────────────────────────
ws1 = wb.create_sheet("Feuil1")
ws1.sheet_view.showGridLines = False

# Title
ws1.merge_cells("A1:J1")
hcell(ws1, 1, 1, "FEUILLE DE TRAVAIL — EXHAUSTIVITÉ DES CHARGES DU PERSONNEL",
      font=F_TITLE, fill=FILL_ALT, align=CTR)
ws1.row_dimensions[1].height = 30

ws1.merge_cells("A2:J2")
hcell(ws1, 2, 1, "Rapprochement Paie / Comptabilité — Salaire Brut et Charges Patronales",
      font=F_SUBHDR, align=CTR)

# Metadata section
meta_fields = [
    ("Société / Company", ""),
    ("Exercice / Period", ""),
    ("Auditeur / Auditor", ""),
    ("Date rapport / Report date", ""),
]
for i, (label, val) in enumerate(meta_fields, 4):
    hcell(ws1, i, 1, label, font=F_SUBHDR, align=LFT, bdr=BDR_DATA)
    ws1.merge_cells(f"B{i}:E{i}")
    hcell(ws1, i, 2, val, font=F_DATA, align=LFT, bdr=BDR_DATA)

ws1.row_dimensions[3].height = 6  # spacer

# Summary table header (row 9)
summary_headers = ["", "SAL BRUT (R)", "CNPS/P (S)", "CF/P (T)", "FNE (U)", "V (T+U)", "AF (W)", "AT (X)", "TOTAL (Y)"]
for c_idx, h in enumerate(summary_headers, 1):
    cell = ws1.cell(9, c_idx, value=h)
    cell.font = F_HEADER; cell.fill = FILL_BLUE
    cell.alignment = CTR; cell.border = BDR_DATA

# PAIE row
ws1.merge_cells("A10:A10")
hcell(ws1, 10, 1, "TOTAL PAIE", font=F_TOT_P, fill=FILL_TOT_P, align=LFT, bdr=BDR_TOT)
for c in range(2, 10):
    hcell(ws1, 10, c, None, font=F_TOT_P, fill=FILL_TOT_P, align=RGT, bdr=BDR_TOT, num_fmt=NUM_FMT)

# COMPTA row
hcell(ws1, 11, 1, "TOTAL COMPTABILITÉ", font=F_TOT_C, fill=FILL_TOT_C, align=LFT, bdr=BDR_TOT)
for c in range(2, 10):
    hcell(ws1, 11, c, None, font=F_TOT_C, fill=FILL_TOT_C, align=RGT, bdr=BDR_TOT, num_fmt=NUM_FMT)

# ECART row
hcell(ws1, 12, 1, "ÉCART", font=F_ECART, fill=FILL_ECART, align=LFT, bdr=BDR_TOT)
for c in range(2, 10):
    hcell(ws1, 12, c, None, font=F_ECART, fill=FILL_ECART, align=RGT, bdr=BDR_TOT, num_fmt=NUM_FMT)

# Observations section
hcell(ws1, 14, 1, "OBSERVATIONS / FINDINGS", font=F_SUBHDR, align=LFT, bdr=BDR_DATA)
ws1.merge_cells("A15:J20")
obs_cell = ws1.cell(15, 1, value="")
obs_cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
obs_cell.border = BDR_HEAVY; obs_cell.font = F_DATA
ws1.row_dimensions[15].height = 80

# Statut
hcell(ws1, 22, 1, "STATUT / STATUS", font=F_SUBHDR, align=LFT)
ws1.merge_cells("B22:E22")
hcell(ws1, 22, 2, "En attente / Pending", font=F_DATA, fill=FILL_YELLOW, align=CTR)

# Column widths Feuil1
ws1.column_dimensions["A"].width = 30
for col in "BCDEFGHIJ":
    ws1.column_dimensions[col].width = 16

ws1.freeze_panes = "A9"

# ─────────────────────────────────────────────────────────────────
# FEUIL2 — Reconciliation Workpaper (Blank Structure)
# ─────────────────────────────────────────────────────────────────
ws2 = wb.create_sheet("Feuil2")
ws2.sheet_view.showGridLines = False

# Title row 1
ws2.merge_cells("A1:Y1")
hcell(ws2, 1, 1, "RAPPROCHEMENT PAIE / COMPTABILITÉ — CHARGES DU PERSONNEL",
      font=F_TITLE, fill=FILL_ALT, align=CTR)
ws2.row_dimensions[1].height = 25

# Section I header (row 2)
ws2.merge_cells("A2:Y2")
hcell(ws2, 2, 1, "I. COÛT TOTAL PAIE",
      font=Font(name="Arial", bold=True, size=11, color="1F4E79"),
      fill=FILL_TOT_P, align=LFT)

# Column headers (row 3)
COL_HEADERS = [
    (1, "Matricule"), (2, "Nom"), (3, "Prénom"), (4, "Service"), (5, "Catégorie"),
    (6, "Col F"), (7, "Col G"), (8, "Col H"), (9, "Col I"), (10, "Col J"),
    (11, "Col K"), (12, "Col L"), (13, "Col M"),
    (14, "Salaire Base"),      # N — kept blank
    (15, "Ancienneté"),        # O — kept blank
    (16, "H. Sup"),            # P — kept blank
    (17, "Autre Gain"),        # Q — kept blank
    (18, "SAL BRUT"),          # R
    (19, "CNPS/P"),            # S
    (20, "CF/P"),              # T
    (21, "FNE"),               # U
    (22, "CF/P+FNE"),          # V
    (23, "AF"),                # W
    (24, "AT"),                # X
    (25, "TOTAL"),             # Y
]
for col_idx, label in COL_HEADERS:
    cell = ws2.cell(3, col_idx, value=label)
    cell.font = F_HEADER; cell.fill = FILL_BLUE
    cell.alignment = CTR; cell.border = BDR_DATA

ws2.row_dimensions[3].height = 40
ws2.freeze_panes = "A4"

# Placeholder note row 4
ws2.merge_cells("A4:M4")
hcell(ws2, 4, 1, "[Lignes employés PAIE — remplies par le plugin / Employee rows filled by plugin]",
      font=Font(name="Arial", italic=True, size=9, color="808080"), align=LFT)

# TOTAL PAIE row (178)
TOTAL_PAIE_ROW = 178
hcell(ws2, TOTAL_PAIE_ROW, 1, "TOTAL PAIE",
      font=F_TOT_P, fill=FILL_TOT_P, align=LFT, bdr=BDR_TOT)
for col in range(18, 26):
    hcell(ws2, TOTAL_PAIE_ROW, col, None,
          font=F_TOT_P, fill=FILL_TOT_P, align=RGT, bdr=BDR_TOT, num_fmt=NUM_FMT)

# Section II header (row 179)
ws2.merge_cells(f"A179:Y179")
hcell(ws2, 179, 1, "II. COÛT TOTAL COMPTABILITÉ (Grand Livre — comptes 66x)",
      font=Font(name="Arial", bold=True, size=11, color="1F4E79"),
      fill=FILL_TOT_C, align=LFT)

# Placeholder for COMPTA rows (row 180)
ws2.merge_cells("A180:M180")
hcell(ws2, 180, 1, "[Lignes comptes GL — remplies par le plugin / GL account rows filled by plugin]",
      font=Font(name="Arial", italic=True, size=9, color="808080"), align=LFT)

# TOTAL COMPTA placeholder (row 194)
TOTAL_COMPTA_ROW = 194
hcell(ws2, TOTAL_COMPTA_ROW, 1, "TOTAL COMPTABILITÉ",
      font=F_TOT_C, fill=FILL_TOT_C, align=LFT, bdr=BDR_TOT)
for col in range(18, 26):
    hcell(ws2, TOTAL_COMPTA_ROW, col, None,
          font=F_TOT_C, fill=FILL_TOT_C, align=RGT, bdr=BDR_TOT, num_fmt=NUM_FMT)

# Section III ECART header (row 195)
ws2.merge_cells("A195:Y195")
hcell(ws2, 195, 1, "III. ÉCARTS (= COMPTABILITÉ − PAIE)",
      font=Font(name="Arial", bold=True, size=11, color="843C0C"),
      fill=FILL_ALT, align=LFT)

# ECART ROW (197)
ECART_ROW = 197
hcell(ws2, ECART_ROW, 1, "ÉCART TOTAL",
      font=F_ECART, fill=FILL_ECART, align=LFT, bdr=BDR_TOT)
for col in range(18, 26):
    hcell(ws2, ECART_ROW, col, None,
          font=F_ECART, fill=FILL_ECART, align=RGT, bdr=BDR_TOT, num_fmt=NUM_FMT)

# Column widths Feuil2
ws2.column_dimensions["A"].width = 12
ws2.column_dimensions["B"].width = 20
ws2.column_dimensions["C"].width = 16
for i in range(4, 14):
    ws2.column_dimensions[gcl(i)].width = 10
for i in range(14, 18):
    ws2.column_dimensions[gcl(i)].width = 14  # N/O/P/Q — kept blank
for i in range(18, 26):
    ws2.column_dimensions[gcl(i)].width = 16  # R–Y

# Row heights
for r in range(4, 178):
    ws2.row_dimensions[r].height = 18
ws2.row_dimensions[TOTAL_PAIE_ROW].height = 22
ws2.row_dimensions[TOTAL_COMPTA_ROW].height = 22
ws2.row_dimensions[ECART_ROW].height = 22

# ─────────────────────────────────────────────────────────────────
# FEUIL3 — Detail / Supporting Schedules
# ─────────────────────────────────────────────────────────────────
ws3 = wb.create_sheet("Feuil3")
ws3.sheet_view.showGridLines = False

ws3.merge_cells("A1:H1")
hcell(ws3, 1, 1, "FEUIL3 — DÉTAIL DES CALCULS / SCHEDULES",
      font=F_TITLE, fill=FILL_ALT, align=CTR)
ws3.row_dimensions[1].height = 25

sections = [
    (3,  "A. Détail des comptes 661-663 (Salaires et appointements)"),
    (20, "B. Détail des comptes 664xxx (Cotisations patronales)"),
    (35, "C. Récapitulatif par département / service"),
    (50, "D. Notes et références"),
]
for row, label in sections:
    ws3.merge_cells(f"A{row}:H{row}")
    hcell(ws3, row, 1, label,
          font=Font(name="Arial", bold=True, size=10, color="1F4E79"),
          fill=FILL_ALT, align=LFT, bdr=BDR_TOT)
    for c in range(1, 9):
        hcell(ws3, row+1, c, None, font=F_DATA, align=LFT, bdr=BDR_DATA)

ws3.column_dimensions["A"].width = 16
ws3.column_dimensions["B"].width = 35
for col in "CDEFGH":
    ws3.column_dimensions[col].width = 18

# ─────────────────────────────────────────────────────────────────
# Save
# ─────────────────────────────────────────────────────────────────
out_dir = os.path.join(os.path.dirname(__file__), "..", "assets")
os.makedirs(out_dir, exist_ok=True)
out_path = os.path.join(out_dir, "FT-P-2-template.xlsx")

wb.save(out_path)
print(f"✅ Template created: {out_path}")
print(f"   Sheets: {wb.sheetnames}")
