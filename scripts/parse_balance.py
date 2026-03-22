"""
parse_balance.py — Parse Balance Générale (xlsx) → 3-part structured Extract Balance sheet.

Part 1: 661x + 663x (Rémunérations directes) — subtotal feeds Feuil2 Section A col R
Part 2: 664110, 664120, 664130 (CNPS) — subtotal feeds Feuil2 Section B cols S,W,X
Part 3: 664380 + FNE note (CF/P+FNE) — subtotal feeds Feuil2 Section C cols T,U

Source columns: col4=MvtDebit, col5=MvtCredit (standard 8-column BG format)
Solde Net = MvtDebit - MvtCredit
"""
import io, json, sys
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

SESSION_FILE = ".audit-session.json"
with open(SESSION_FILE, encoding="utf-8") as f:
    session = json.load(f)

BG_PATH = session["files"]["balance_generale"]
FT_PATH = session["files"]["feuille_travail"]
ACCT_PLAN = session.get("accounting_plan", "SYSCOHADA")

print(f"Parsing Balance Générale: {BG_PATH}")
print(f"Accounting plan: {ACCT_PLAN}")

# ── Read BG ───────────────────────────────────────────────────────────────────
df = pd.read_excel(BG_PATH, dtype={0: str})
df.columns = [str(c).strip() for c in df.columns]
print(f"  Columns ({len(df.columns)}): {list(df.columns)}")

# Standard column positions (index-based — more robust than name matching)
compte_col = df.columns[0]
libelle_col = df.columns[1]
mvtd_col = df.columns[4] if len(df.columns) > 4 else None
mvtc_col = df.columns[5] if len(df.columns) > 5 else None

if not mvtd_col or not mvtc_col:
    print(f"❌ BG file has fewer than 6 columns — cannot read MvtDebit/MvtCredit")
    sys.exit(1)

df[compte_col] = df[compte_col].astype(str).str.strip().str.zfill(0)
df["MvtDebit"]  = pd.to_numeric(df[mvtd_col], errors="coerce").fillna(0)
df["MvtCredit"] = pd.to_numeric(df[mvtc_col], errors="coerce").fillna(0)
df["SoldeNet"]  = df["MvtDebit"] - df["MvtCredit"]

print(f"  Total BG rows: {len(df)}")

# ── Account group definitions (SYSCOHADA) ─────────────────────────────────────
def filter_accounts(df, prefixes):
    mask = df[compte_col].str.match('|'.join(f'^{p}' for p in prefixes))
    return df[mask & (df["SoldeNet"].abs() > 0)].copy()

PART1_PREFIXES = ["661", "663"]
PART2_ACCOUNTS = ["664110", "664120", "664130"]
PART3_ACCOUNTS = ["664380"]

df_p1 = filter_accounts(df, PART1_PREFIXES).sort_values(compte_col)
df_p2 = df[df[compte_col].isin(PART2_ACCOUNTS)].copy().sort_values(compte_col)
df_p3 = df[df[compte_col].isin(PART3_ACCOUNTS)].copy()

print(f"  Part 1 (661x+663x): {len(df_p1)} accounts | "
      f"SoldeNet = {df_p1['SoldeNet'].sum():,.0f} FCFA")
print(f"  Part 2 (CNPS 664110/120/130): {len(df_p2)} accounts")
print(f"  Part 3 (CF/P 664380): {len(df_p3)} accounts")

# ── Styling ───────────────────────────────────────────────────────────────────
BLUE_FILL  = PatternFill("solid", start_color="1F4E79")
WHT_FONT   = Font(name="Arial", bold=True, color="FFFFFF", size=10)
DAT_FONT   = Font(name="Arial", size=10)
GRP_FONT   = Font(name="Arial", bold=True, size=10, color="1F4E79")
GRP_FILL   = PatternFill("solid", start_color="D6E4F0")
STOT_FONT  = Font(name="Arial", bold=True, size=10, color="FFFFFF")
STOT_FILL  = PatternFill("solid", start_color="2E75B6")
ALT_FILL   = PatternFill("solid", start_color="F5F8FC")
FNE_FILL   = PatternFill("solid", start_color="FFF2CC")  # yellow for FNE note
FNE_FONT   = Font(name="Arial", italic=True, size=10, color="806000")
RGT = Alignment(horizontal="right", vertical="center")
LFT = Alignment(horizontal="left",  vertical="center")
NUM_FMT = "#,##0;(#,##0);\"-\""

def thn(): return Side(style="thin",   color="D9D9D9")
def med(): return Side(style="medium", color="1F4E79")
BDR   = Border(bottom=thn(), right=thn())
BDR_T = Border(top=med(), bottom=med(), right=thn())

# ── Write to workbook ─────────────────────────────────────────────────────────
wb = load_workbook(FT_PATH)
if "Extract Balance" in wb.sheetnames:
    del wb["Extract Balance"]
ws = wb.create_sheet("Extract Balance")

current_row = [1]  # mutable reference

def write_header():
    headers = ["Compte", "Libellé", "Mvt Débit", "Mvt Crédit", "Solde Net"]
    for c, h in enumerate(headers, 1):
        cell = ws.cell(current_row[0], c, value=h)
        cell.font = WHT_FONT; cell.fill = BLUE_FILL
        cell.alignment = LFT; cell.border = BDR
    current_row[0] += 1

def write_section_label(label):
    ws.merge_cells(f"A{current_row[0]}:E{current_row[0]}")
    cell = ws.cell(current_row[0], 1, value=label)
    cell.font = GRP_FONT; cell.fill = GRP_FILL; cell.alignment = LFT; cell.border = BDR_T
    current_row[0] += 1

def write_data_row(compte, libelle, mvt_d, mvt_c, solde, fill=None):
    vals = [compte, libelle, mvt_d, mvt_c, solde]
    for c, val in enumerate(vals, 1):
        cell = ws.cell(current_row[0], c, value=val)
        cell.font = DAT_FONT; cell.border = BDR
        if fill: cell.fill = fill
        cell.alignment = RGT if c >= 3 else LFT
        if c >= 3: cell.number_format = NUM_FMT
    current_row[0] += 1

def write_subtotal(label, mvt_d, mvt_c, solde):
    vals = [label, None, mvt_d, mvt_c, solde]
    for c, val in enumerate(vals, 1):
        cell = ws.cell(current_row[0], c, value=val)
        cell.font = STOT_FONT; cell.fill = STOT_FILL; cell.border = BDR_T
        cell.alignment = RGT if c >= 3 else LFT
        if c >= 3: cell.number_format = NUM_FMT
    current_row[0] += 1

write_header()

# ── Part 1 — Rémunérations directes (661x + 663x) ────────────────────────────
write_section_label("Partie 1 — Rémunérations directes et avantages (661x + 663x)")
for i, row in enumerate(df_p1.itertuples(index=False)):
    alt = ALT_FILL if (i % 2 == 0) else None
    write_data_row(
        getattr(row, compte_col), getattr(row, libelle_col),
        row.MvtDebit, row.MvtCredit, row.SoldeNet, fill=alt
    )
p1_d, p1_c, p1_s = df_p1["MvtDebit"].sum(), df_p1["MvtCredit"].sum(), df_p1["SoldeNet"].sum()
write_subtotal("Sous-total 661-663", p1_d, p1_c, p1_s)
current_row[0] += 1  # blank spacer

# ── Part 2 — Cotisations CNPS (664110, 664120, 664130) ───────────────────────
write_section_label("Partie 2 — Cotisations CNPS (664110 AF | 664120 CNPS/P | 664130 AT)")
for i, row in enumerate(df_p2.itertuples(index=False)):
    alt = ALT_FILL if (i % 2 == 0) else None
    write_data_row(
        getattr(row, compte_col), getattr(row, libelle_col),
        row.MvtDebit, row.MvtCredit, row.SoldeNet, fill=alt
    )
p2_d, p2_c, p2_s = df_p2["MvtDebit"].sum(), df_p2["MvtCredit"].sum(), df_p2["SoldeNet"].sum()
write_subtotal("Sous-total CNPS", p2_d, p2_c, p2_s)
current_row[0] += 1

# ── Part 3 — CF/P et FNE (664380 + note FNE) ─────────────────────────────────
write_section_label("Partie 3 — Crédit Foncier Patronal & FNE (664380 + FNE hors GL)")
for i, row in enumerate(df_p3.itertuples(index=False)):
    alt = ALT_FILL if (i % 2 == 0) else None
    write_data_row(
        getattr(row, compte_col), getattr(row, libelle_col),
        row.MvtDebit, row.MvtCredit, row.SoldeNet, fill=alt
    )
# FNE note row
write_data_row("—", "FNE — non comptabilisé en GL (retenue salariale hors 66x)", 0, 0, 0, fill=FNE_FILL)
# Style the FNE note differently
fne_row = current_row[0] - 1
for c in range(1, 6):
    ws.cell(fne_row, c).font = FNE_FONT

p3_d, p3_c, p3_s = df_p3["MvtDebit"].sum(), df_p3["MvtCredit"].sum(), df_p3["SoldeNet"].sum()
write_subtotal("Sous-total CF/P+FNE", p3_d, p3_c, p3_s)

# Column widths
ws.column_dimensions["A"].width = 12
ws.column_dimensions["B"].width = 45
ws.column_dimensions["C"].width = 20
ws.column_dimensions["D"].width = 20
ws.column_dimensions["E"].width = 20

ws.auto_filter.ref = "A1:E1"
ws.freeze_panes = "A2"

wb.save(FT_PATH)
print(f"✅ Extract Balance: Part1={len(df_p1)} rows (sous-total {p1_s:,.0f}) | "
      f"Part2={len(df_p2)} rows (sous-total {p2_s:,.0f}) | "
      f"Part3={len(df_p3)} rows (sous-total {p3_s:,.0f}) → '{FT_PATH}'")

# Store BG account values in session for reconciliation
session.setdefault("bg_amounts", {})
for row in df_p1.itertuples(index=False):
    session["bg_amounts"][getattr(row, compte_col)] = round(row.SoldeNet)
for row in df_p2.itertuples(index=False):
    session["bg_amounts"][getattr(row, compte_col)] = round(row.SoldeNet)
for row in df_p3.itertuples(index=False):
    session["bg_amounts"][getattr(row, compte_col)] = round(row.SoldeNet)
session["bg_totals"] = {
    "part1_solde": round(p1_s), "part2_solde": round(p2_s), "part3_solde": round(p3_s)
}

session.setdefault("extract_counts", {})["balance"] = len(df_p1) + len(df_p2) + len(df_p3)
session.setdefault("steps_completed", [])
if "extract_balance" not in session["steps_completed"]:
    session["steps_completed"].append("extract_balance")
with open(SESSION_FILE, "w", encoding="utf-8") as f:
    json.dump(session, f, indent=2, ensure_ascii=False)
