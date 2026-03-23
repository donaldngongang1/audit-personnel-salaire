"""
build_reconciliation.py — Build Feuil2 (v1.2 — column-name-driven formulas).

All formulas and column references are derived from the Feuil2 header row at runtime
via col_utils.get_feuil2_col_map(). If columns move, formulas stay correct.

Section A source: Balance Générale (BG xlsx) — col4=MvtDebit, col5=MvtCredit
Section B/C/D source: Grand Livre (GL xls) — ALL journals

Usage: python build_reconciliation.py [--section paie|compta|all]
"""
import argparse, io, json, subprocess, sys
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

import pandas as pd
import xlrd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter as gcl

from col_utils import (
    get_feuil2_col_map, describe_col_map,
    all_numeric_cols,
    build_v_formula, build_y_formula,
    build_sum_formula, build_total_compta_formula, build_ecart_formula,
)

# ── Load session ───────────────────────────────────────────────────────────────
SESSION_FILE = ".audit-session.json"
with open(SESSION_FILE, encoding="utf-8") as f:
    session = json.load(f)

FT_PATH = session["files"]["feuille_travail"]
BG_PATH = session["files"]["balance_generale"]
GL_PATH = session["files"]["grand_livre"]

# ── Styling constants ──────────────────────────────────────────────────────────
NUM_FMT = "#,##0;(#,##0);\"-\""

def med(c="1F4E79"): return Side(style="medium", color=c)
def thn(c="D9D9D9"): return Side(style="thin",   color=c)

BDR_DATA = Border(bottom=thn(), right=thn())
BDR_TOT  = Border(top=med(), bottom=med(), right=thn())

RGT = Alignment(horizontal="right", vertical="center")
LFT = Alignment(horizontal="left",  vertical="center")

FDATA   = Font(name="Arial", size=10)
FWHITE  = Font(name="Arial", bold=True, color="FFFFFF", size=10)
FTOT_P  = Font(name="Arial", bold=True, size=10, color="1F4E79")
FTOT_C  = Font(name="Arial", bold=True, size=10, color="1F4E79")
FGRP    = Font(name="Arial", bold=True, size=10, color="1F4E79")

FILL_ALT   = PatternFill("solid", start_color="F5F8FC")
FILL_TOT_P = PatternFill("solid", start_color="D6E4F0")
FILL_TOT_C = PatternFill("solid", start_color="C6EFCE")
FILL_ECART = PatternFill("solid", start_color="843C0C")
FILL_GRP   = PatternFill("solid", start_color="EBF3FB")
FILL_STOT  = PatternFill("solid", start_color="2E75B6")
FILL_NONE  = PatternFill(fill_type=None)
FILL_D     = PatternFill("solid", start_color="808080")

# ── Cell helpers ───────────────────────────────────────────────────────────────
def blank_cell(ws, row, col, fill=None):
    c = ws.cell(row, col)
    c.value = None; c.font = FDATA; c.number_format = "General"
    c.alignment = RGT; c.border = BDR_DATA
    c.fill = fill if fill else FILL_NONE

def num_cell(ws, row, col, value, font, fill, bdr=BDR_DATA):
    c = ws.cell(row, col, value=value)
    c.font = font; c.fill = fill if fill else FILL_NONE
    c.number_format = NUM_FMT; c.alignment = RGT; c.border = bdr

def formula_cell(ws, row, col, formula, font, fill, bdr=BDR_DATA):
    c = ws.cell(row, col, value=formula)
    c.font = font; c.fill = fill if fill else FILL_NONE
    c.number_format = NUM_FMT; c.alignment = RGT; c.border = bdr

def label_cell(ws, row, col, value, font, fill, bdr=BDR_DATA):
    c = ws.cell(row, col, value=value)
    c.font = font; c.fill = fill if fill else FILL_NONE
    c.alignment = LFT; c.border = bdr

# ── Locate TOTAL PAIE row ──────────────────────────────────────────────────────
def find_total_paie_row(ws):
    for r in range(1, ws.max_row + 1):
        v = str(ws.cell(r, 1).value or ws.cell(r, 2).value or "")
        if "TOTAL PAIE" in v.upper() and "COMPTA" not in v.upper():
            return r
    return 178  # fallback

# ── Load payroll pivot ─────────────────────────────────────────────────────────
def load_paie_data():
    lp = pd.read_csv(".livre_paie_pivot.csv", dtype={"Matricule": str})
    cp = pd.read_csv(".charges_patronales_pivot.csv", dtype={"Matricule": str})
    COL_MAP = {
        "Credit_Foncier_Patronal":      "CF_P",
        "Fond_National_de_lemploi_FNE": "FNE",
        "Pension_Vieillesse_CNPS":      "CNPS_P",
        "Allocation_Familiale":         "AF",
        "Accident_de_Travail":          "AT",
    }
    cp = cp.rename(columns={k: v for k, v in COL_MAP.items() if k in cp.columns})
    merged = pd.merge(lp, cp, on=["Matricule", "Nom", "Prenom"], how="outer").fillna(0)
    for col in ["SAL_BRUT", "CF_P", "FNE", "CNPS_P", "AF", "AT"]:
        if col not in merged.columns:
            merged[col] = 0
    return merged.sort_values("Matricule").reset_index(drop=True)

# ── Load BG (Section A: 661x+663x) ────────────────────────────────────────────
def load_bg_section_a():
    df = pd.read_excel(BG_PATH, dtype={0: str})
    df.columns = [str(c).strip() for c in df.columns]
    compte_col = df.columns[0]
    df[compte_col] = df[compte_col].astype(str).str.strip()
    df["MvtDebit"]  = pd.to_numeric(df.iloc[:, 4], errors="coerce").fillna(0)
    df["MvtCredit"] = pd.to_numeric(df.iloc[:, 5], errors="coerce").fillna(0)
    df["SoldeNet"]  = df["MvtDebit"] - df["MvtCredit"]
    mask = df[compte_col].str.match(r'^661|^663')
    df_a = df[mask & (df["SoldeNet"].abs() > 0)].copy()
    return {str(row[compte_col]): {"libelle": str(row.iloc[1]), "solde": round(row["SoldeNet"])}
            for _, row in df_a.iterrows()}

# ── Load GL (Sections B, C, D: 664x+668x, all journals) ───────────────────────
def load_gl_sections():
    xls = xlrd.open_workbook(GL_PATH)
    sheet_name = "Sage" if "Sage" in xls.sheet_names() else xls.sheet_names()[0]
    df = pd.read_excel(GL_PATH, sheet_name=sheet_name, engine="xlrd", dtype={0: str})
    df.columns = [str(c).strip() for c in df.columns]
    compte_col  = df.columns[0]
    debit_col   = next((c for c in df.columns if "debit"  in c.lower()), df.columns[8] if len(df.columns) > 8 else None)
    credit_col  = next((c for c in df.columns if "credit" in c.lower()), df.columns[9] if len(df.columns) > 9 else None)
    libelle_col = df.columns[4] if len(df.columns) > 4 else df.columns[1]
    df[compte_col] = df[compte_col].astype(str).str.strip()
    df["Debit"]  = pd.to_numeric(df[debit_col],  errors="coerce").fillna(0) if debit_col  else 0
    df["Credit"] = pd.to_numeric(df[credit_col], errors="coerce").fillna(0) if credit_col else 0
    df_6xx = df[df[compte_col].str.match(r'^664|^668')].copy()
    agg = df_6xx.groupby(compte_col).agg(
        Debit=("Debit", "sum"), Credit=("Credit", "sum"), Libelle=(libelle_col, "first")
    ).reset_index()
    agg["SoldeNet"] = agg["Debit"] - agg["Credit"]
    return {str(row[compte_col]): {"libelle": str(row["Libelle"]), "solde": round(row["SoldeNet"])}
            for _, row in agg.iterrows()}

# ── PAIE section ───────────────────────────────────────────────────────────────
def build_paie_section(ws, col_map, paie_df, data_start, tot_paie_row):
    paie_end  = tot_paie_row - 1
    blank_cols = col_map["BLANK_COLS"]
    print(f"Building PAIE section: rows {data_start}–{paie_end} ({len(paie_df)} employees)")
    print(describe_col_map(col_map))

    paie_col_keys = [
        ("SAL_BRUT", "SAL_BRUT"),
        ("CNPS_P",   "CNPS_P"),
        ("CF_P",     "CF_P"),
        ("FNE",      "FNE"),
        ("AF",       "AF"),
        ("AT",       "AT"),
    ]

    for i, emp in enumerate(paie_df.itertuples(index=False)):
        r = data_start + i
        if r >= tot_paie_row:
            print(f"⚠️ More employees than PAIE rows — stopping at row {r-1}")
            break
        alt = FILL_ALT if (i % 2 == 0) else None

        for col in blank_cols:
            blank_cell(ws, r, col, fill=alt)

        # Write numeric columns — look up each col index by logical name
        for logical_name, attr_name in paie_col_keys:
            col_idx = col_map[logical_name]
            val     = round(getattr(emp, attr_name, 0) or 0)
            num_cell(ws, r, col_idx, val, FDATA, alt)

        # Derived formulas — built from col names, not hardcoded letters
        cf_fne_idx = col_map["CF_FNE"]
        total_idx  = col_map["TOTAL_COL"]
        formula_cell(ws, r, cf_fne_idx, build_v_formula(col_map, r), FDATA, alt)
        formula_cell(ws, r, total_idx,  build_y_formula(col_map, r), FDATA, alt)

    # TOTAL PAIE row
    r = tot_paie_row
    for col in blank_cols:
        blank_cell(ws, r, col, fill=FILL_TOT_P)

    for logical in ["SAL_BRUT", "CNPS_P", "CF_P", "FNE", "CF_FNE", "AF", "AT", "TOTAL_COL"]:
        formula_cell(ws, r, col_map[logical],
                     build_sum_formula(col_map, logical, data_start, paie_end),
                     FTOT_P, FILL_TOT_P, BDR_TOT)

    print(f"  TOTAL PAIE at row {r}: SUM over rows {data_start}:{paie_end}")


# ── Write helpers for COMPTA ───────────────────────────────────────────────────
def write_group_label(ws, row, label, n_cols):
    for col in range(1, n_cols + 1):
        c = ws.cell(row, col)
        c.value = None; c.fill = FILL_GRP; c.border = BDR_TOT; c.font = FGRP
    ws.cell(row, 1, value=label).alignment = LFT

def write_account_row(ws, row, col_map, compte, libelle, logical_target, amount, fill=None):
    for col in col_map["BLANK_COLS"]:
        blank_cell(ws, row, col, fill=fill)
    label_cell(ws, row, 1, compte,  FDATA, fill)
    label_cell(ws, row, 2, libelle, FDATA, fill)
    if logical_target:
        num_cell(ws, row, col_map[logical_target], amount, FDATA, fill)
    formula_cell(ws, row, col_map["CF_FNE"],    build_v_formula(col_map, row), FDATA, fill)
    formula_cell(ws, row, col_map["TOTAL_COL"], build_y_formula(col_map, row), FDATA, fill)

def write_subtotal_row(ws, row, col_map, label, group_start, group_end, fill=FILL_STOT, font=FWHITE):
    for col in col_map["BLANK_COLS"]:
        blank_cell(ws, row, col, fill=fill)
    label_cell(ws, row, 1, label, font, fill, BDR_TOT)
    for logical in ["SAL_BRUT", "CNPS_P", "CF_P", "FNE", "CF_FNE", "AF", "AT", "TOTAL_COL"]:
        formula_cell(ws, row, col_map[logical],
                     build_sum_formula(col_map, logical, group_start, group_end),
                     font, fill, BDR_TOT)

# ── COMPTA section ─────────────────────────────────────────────────────────────
def build_compta_section(ws, col_map, bg_map, gl_map, start_row):
    r = start_row
    n_cols = max(all_numeric_cols(col_map)) + 2
    subtotal_rows = []

    # ── Group A — 661x+663x (source: BG) ──────────────────────────────────────
    write_group_label(ws, r, "Section A — Rémunérations directes (661x+663x) — Source: Balance Générale", n_cols)
    r += 1; group_a_start = r
    A_ACCOUNTS = ["661110","661120","661130","661200","661210","661220",
                  "661300","661380","661410","661800","663101","663102","663410"]
    for i, acct in enumerate(A_ACCOUNTS):
        info = bg_map.get(acct, {"libelle": acct, "solde": 0})
        write_account_row(ws, r, col_map, acct, info["libelle"], "SAL_BRUT", info["solde"],
                          fill=FILL_ALT if i % 2 == 0 else None)
        r += 1
    write_subtotal_row(ws, r, col_map, "Sous-total A — 661x+663x", group_a_start, r - 1)
    subtotal_rows.append(r)
    print(f"  Group A: rows {group_a_start}–{r-1}, subtotal {r}")
    r += 2

    # ── Group B — CNPS 664110/120/130 (source: GL) ────────────────────────────
    write_group_label(ws, r, "Section B — Cotisations CNPS (664110 AF | 664120 CNPS/P | 664130 AT) — Source: Grand Livre", n_cols)
    r += 1; group_b_start = r
    B_ACCOUNTS = [
        ("664120", "CNPS Pension Vieillesse (AV)", "CNPS_P"),
        ("664110", "CNPS Allocation Familiale (AF)", "AF"),
        ("664130", "CNPS Accident de Travail (AT)",  "AT"),
    ]
    for i, (acct, fallback, logical) in enumerate(B_ACCOUNTS):
        info = gl_map.get(acct, {"libelle": fallback, "solde": 0})
        write_account_row(ws, r, col_map, acct, info["libelle"], logical, info["solde"],
                          fill=FILL_ALT if i % 2 == 0 else None)
        r += 1
    write_subtotal_row(ws, r, col_map, "Sous-total B — CNPS", group_b_start, r - 1)
    subtotal_rows.append(r)
    print(f"  Group B: rows {group_b_start}–{r-1}, subtotal {r}")
    r += 2

    # ── Group C — CF/P + FNE (source: GL) ─────────────────────────────────────
    write_group_label(ws, r, "Section C — Crédit Foncier Patronal & FNE (664380 + FNE=0) — Source: Grand Livre", n_cols)
    r += 1; group_c_start = r
    cfp_info = gl_map.get("664380", {"libelle": "Provisions Crédit Foncier Patronal", "solde": 0})
    write_account_row(ws, r, col_map, "664380", cfp_info["libelle"], "CF_P", cfp_info["solde"],
                      fill=FILL_ALT)
    r += 1
    write_account_row(ws, r, col_map, "—", "FNE — non comptabilisé en GL (retenue salariale hors 66x)",
                      "FNE", 0, fill=None)
    ws.cell(r, 2).font = Font(name="Arial", italic=True, size=10, color="806000")
    r += 1
    write_subtotal_row(ws, r, col_map, "Sous-total C — CF/P+FNE", group_c_start, r - 1)
    subtotal_rows.append(r)
    print(f"  Group C: rows {group_c_start}–{r-1}, subtotal {r}")
    r += 2

    # ── Group D — 668x (informative only, excluded from TOTAL) ────────────────
    write_group_label(ws, r, "Section D — Autres charges sociales (668x) — Informatif uniquement — NON inclus dans TOTAL", n_cols)
    r += 1; group_d_start = r
    for i, acct in enumerate(["668420", "668430", "668700"]):
        info = gl_map.get(acct, {"libelle": acct, "solde": 0})
        write_account_row(ws, r, col_map, acct, info["libelle"], "SAL_BRUT", info["solde"],
                          fill=FILL_ALT if i % 2 == 0 else None)
        r += 1
    write_subtotal_row(ws, r, col_map, "Sous-total D — 668x (hors rapprochement)",
                       group_d_start, r - 1, fill=FILL_D, font=FWHITE)
    print(f"  Group D (info): rows {group_d_start}–{r-1}, subtotal {r}")
    r += 2

    # ── TOTAL COMPTABILITE = A + B + C (D excluded) ───────────────────────────
    tot_compta_row = r
    for col in col_map["BLANK_COLS"]:
        blank_cell(ws, tot_compta_row, col, fill=FILL_TOT_C)
    label_cell(ws, tot_compta_row, 1, "TOTAL COMPTABILITE (A+B+C)", FTOT_C, FILL_TOT_C, BDR_TOT)
    for logical in ["SAL_BRUT", "CNPS_P", "CF_P", "FNE", "CF_FNE", "AF", "AT", "TOTAL_COL"]:
        formula_cell(ws, tot_compta_row, col_map[logical],
                     build_total_compta_formula(col_map, logical, subtotal_rows),
                     FTOT_C, FILL_TOT_C, BDR_TOT)

    print(f"  TOTAL COMPTA: row {tot_compta_row} = subtotals {subtotal_rows}")
    return tot_compta_row

# ── ECART row ─────────────────────────────────────────────────────────────────
def build_ecart_row(ws, col_map, tot_compta_row, tot_paie_row):
    r = tot_compta_row + 3
    n_cols = max(all_numeric_cols(col_map)) + 2
    for col in range(1, n_cols + 1):
        ws.cell(r, col).fill = FILL_ECART; ws.cell(r, col).border = BDR_TOT
        ws.cell(r, col).value = None; ws.cell(r, col).font = FWHITE
    for col in col_map["BLANK_COLS"]:
        blank_cell(ws, r, col, fill=FILL_ECART)
    label_cell(ws, r, 1, "ECART TOTAL (COMPTABILITE - PAIE)", FWHITE, FILL_ECART, BDR_TOT)
    for logical in ["SAL_BRUT", "CNPS_P", "CF_P", "FNE", "CF_FNE", "AF", "AT", "TOTAL_COL"]:
        formula_cell(ws, r, col_map[logical],
                     build_ecart_formula(col_map, logical, tot_compta_row, tot_paie_row),
                     FWHITE, FILL_ECART, BDR_TOT)
    print(f"  ECART row: {r}")
    return r

# ── Main ──────────────────────────────────────────────────────────────────────
def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--section", default="all", choices=["paie", "compta", "all"])
    args = parser.parse_args()

    wb = load_workbook(FT_PATH)
    ws = wb["Feuil2"]

    # Read column map from Feuil2 header row — everything is driven by this
    col_map = get_feuil2_col_map(ws, header_row=3)

    tot_paie_row = find_total_paie_row(ws)
    compta_start = tot_paie_row + 3

    # Unmerge COMPTA region
    to_unmerge = [mr for mr in list(ws.merged_cells.ranges) if mr.min_row >= compta_start]
    for mr in to_unmerge:
        ws.unmerge_cells(str(mr))

    tot_compta_row = None
    ecart_row = None

    if args.section in ("paie", "all"):
        paie_df = load_paie_data()
        build_paie_section(ws, col_map, paie_df, data_start=4, tot_paie_row=tot_paie_row)

    if args.section in ("compta", "all"):
        bg_map = load_bg_section_a()
        gl_map = load_gl_sections()
        tot_compta_row = build_compta_section(ws, col_map, bg_map, gl_map, compta_start)

    if args.section == "all" and tot_compta_row:
        ecart_row = build_ecart_row(ws, col_map, tot_compta_row, tot_paie_row)

    wb.save(FT_PATH)
    print(f"\n✅ Feuil2 saved: {FT_PATH}")

    # Persist key row numbers + col map to session
    session.setdefault("steps_completed", [])
    if "reconcile" not in session["steps_completed"]:
        session["steps_completed"].append("reconcile")
    session["feuil2_build"] = {
        "row_total_paie":   tot_paie_row,
        "row_total_compta": tot_compta_row,
        "row_ecart":        ecart_row,
        "col_map": {k: v for k, v in col_map.items() if k != "BLANK_COLS"},
        "blank_cols": col_map.get("BLANK_COLS", []),
        "source_section_A": "BG",
        "source_sections_BCD": "GL_all_journals",
    }
    with open(SESSION_FILE, "w", encoding="utf-8") as f:
        json.dump(session, f, indent=2, ensure_ascii=False)

    # Auto-trigger gap analysis (Rule 6)
    if args.section == "all":
        run_mode = session.get("run_mode", "interactive")
        print(f"\n>>> Triggering gap analysis (mode: {run_mode})...")
        result = subprocess.run(
            [sys.executable, "scripts/build_feuil1_summary.py", "--write",
             "--row-paie",   str(tot_paie_row),
             "--row-compta", str(tot_compta_row or ""),
             "--row-ecart",  str(ecart_row or ""),
             "--mode",       run_mode],
            check=False
        )
        if result.returncode != 0:
            print("⚠️ Gap analysis returned non-zero — check output above")
        else:
            print("✅ Gap analysis and Feuil1 complete")

if __name__ == "__main__":
    main()
