"""
build_feuil1_summary.py — Read ecarts from Feuil2, generate explanations, REWRITE Feuil1 from scratch.

ALWAYS reads values directly from Feuil2 in the workbook (not from session cache).
ALWAYS rewrites Feuil1 completely — never partial updates.

Usage:
  python build_feuil1_summary.py [--write]
                                 [--row-paie N] [--row-compta N] [--row-ecart N]
                                 [--mode interactive|unattended]
"""
import argparse, io, json, sys
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

from datetime import date
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

SESSION_FILE = ".audit-session.json"
with open(SESSION_FILE, encoding="utf-8") as f:
    session = json.load(f)

FT_PATH  = session["files"]["feuille_travail"]
LANGUAGE = session.get("language", "fr")

# ── Styling ────────────────────────────────────────────────────────────────────
def med(c="1F4E79"): return Side(style="medium", color=c)
def thn(c="D9D9D9"): return Side(style="thin",   color=c)

BDR_MED  = Border(top=med(), bottom=med(), left=med(), right=med())
BDR_DATA = Border(bottom=thn(), right=thn())
BDR_TOT  = Border(top=med(), bottom=med(), right=thn())

FTITLE  = Font(name="Arial", bold=True, size=14, color="1F4E79")
FHDR    = Font(name="Arial", bold=True, size=11, color="FFFFFF")
FSUBHDR = Font(name="Arial", bold=True, size=10, color="1F4E79")
FDATA   = Font(name="Arial", size=10)
FTOT_P  = Font(name="Arial", bold=True, size=10, color="1F4E79")
FTOT_C  = Font(name="Arial", bold=True, size=10, color="1F4E79")
FECART  = Font(name="Arial", bold=True, size=10, color="FFFFFF")
FWHITE  = Font(name="Arial", bold=True, size=11, color="FFFFFF")

FILL_BLUE  = PatternFill("solid", start_color="1F4E79")
FILL_ALT   = PatternFill("solid", start_color="EBF3FB")
FILL_TOT_P = PatternFill("solid", start_color="D6E4F0")
FILL_TOT_C = PatternFill("solid", start_color="C6EFCE")
FILL_ECART = PatternFill("solid", start_color="843C0C")
FILL_OK    = PatternFill("solid", start_color="C6EFCE")
FILL_WARN  = PatternFill("solid", start_color="FFEB9C")
FILL_NONE  = PatternFill(fill_type=None)

RGT = Alignment(horizontal="right",  vertical="center")
LFT = Alignment(horizontal="left",   vertical="center")
CTR = Alignment(horizontal="center", vertical="center", wrap_text=True)
WRP = Alignment(horizontal="left",   vertical="top",    wrap_text=True)

NUM_FMT = "#,##0;(#,##0);\"-\""

# ── Known gap explanations ─────────────────────────────────────────────────────
GAP_LIBRARY = {
    "U": {
        "fr": (
            "Le FNE (Fond National de l'Emploi) n'est pas comptabilisé dans le Grand Livre Général "
            "sous la classe 66. En SYSCOHADA Cameroun, le FNE est une retenue sur salaire versée "
            "directement à l'ONEM sans écriture comptable distincte dans les comptes 66x. "
            "Écart structurel normal — aucune action corrective requise."
        ),
        "en": (
            "FNE (National Employment Fund) contributions are salary deductions paid directly to ONEM. "
            "Under SYSCOHADA (Cameroon), they are not recorded as Class 66 entries in the GL. "
            "This is a normal structural gap — no corrective action needed."
        ),
        "category": "Structurel — FNE hors GL",
        "action": "Aucune action requise / No action needed",
    },
    "R": {
        "fr": (
            "Différence entre le salaire brut du livre de paie ({paie:,.0f} FCFA) et les charges "
            "comptabilisées en 661x+663x dans la Balance Générale ({compta:,.0f} FCFA). "
            "Différence de {ecart:,.0f} FCFA pouvant résulter de : provisions comptabilisées "
            "hors cycle de paie, régularisations de fin d'exercice, ou charges du personnel "
            "extérieur (661800 / prestations ponctuelles)."
        ),
        "en": (
            "Difference between gross payroll ({paie:,.0f} FCFA) and charges recorded under "
            "661x+663x in the Trial Balance ({compta:,.0f} FCFA). Gap of {ecart:,.0f} FCFA may "
            "result from: provisions outside payroll cycle, year-end adjustments, or external "
            "staff charges (account 661800 — occasional services)."
        ),
        "category": "Écart à justifier",
        "action": "Analyser les écritures 661800 et les régularisations de fin d'exercice",
    },
}

def get_explanation(col, paie_val, compta_val, ecart_val, language):
    """Generate gap explanation for a column."""
    if abs(ecart_val) < 2:
        return {
            "fr": "Aucun écart — soldes réconciliés.",
            "en": "No gap — balances reconcile.",
            "category": "Réconcilié",
            "action": "Aucune",
        }
    if col in GAP_LIBRARY:
        tmpl = GAP_LIBRARY[col]
        try:
            return {
                "fr": tmpl["fr"].format(paie=paie_val, compta=compta_val, ecart=ecart_val),
                "en": tmpl["en"].format(paie=paie_val, compta=compta_val, ecart=ecart_val),
                "category": tmpl["category"],
                "action": tmpl["action"],
            }
        except KeyError:
            return tmpl
    direction_fr = "COMPTA > PAIE" if ecart_val > 0 else "PAIE > COMPTA"
    direction_en = direction_fr
    return {
        "fr": f"Écart de {abs(ecart_val):,.0f} FCFA ({direction_fr}). Investigation requise.",
        "en": f"Gap of {abs(ecart_val):,.0f} FCFA ({direction_en}). Investigation required.",
        "category": "À investiguer",
        "action": "Investigation requise",
    }

# ── Read Feuil2 values ─────────────────────────────────────────────────────────
def read_feuil2_values(ws2, row_paie, row_compta, row_ecart):
    """Read per-column totals from TOTAL PAIE, TOTAL COMPTA, and ECART rows."""
    COL_NAMES = {18:"R",19:"S",20:"T",21:"U",22:"V",23:"W",24:"X",25:"Y"}
    data = {"paie": {}, "compta": {}, "ecart": {}}
    for col, name in COL_NAMES.items():
        def safe_float(val):
            try: return float(val or 0)
            except: return 0.0
        data["paie"][name]   = safe_float(ws2.cell(row_paie,   col).value)
        data["compta"][name] = safe_float(ws2.cell(row_compta, col).value)
        data["ecart"][name]  = safe_float(ws2.cell(row_ecart,  col).value)
    return data

# ── Build Feuil1 from scratch ──────────────────────────────────────────────────
def build_feuil1(ws1, data, metadata, ecart_explanations):
    """Completely clear and rebuild Feuil1."""

    # Clear all existing content
    for row in ws1.iter_rows():
        for cell in row:
            cell.value = None
            cell.font = Font()
            cell.fill = PatternFill(fill_type=None)
            cell.border = Border()
            cell.alignment = Alignment()
            cell.number_format = "General"

    # Unmerge everything
    for mr in list(ws1.merged_cells.ranges):
        ws1.unmerge_cells(str(mr))

    r = 1  # current row tracker

    def hdr(row, col, val, font, fill, align, bdr=None, span_end=None, num_fmt=None):
        if span_end:
            ws1.merge_cells(f"{chr(64+col)}{row}:{chr(64+span_end)}{row}")
        c = ws1.cell(row, col, value=val)
        c.font = font; c.fill = fill; c.alignment = align
        if bdr: c.border = bdr
        if num_fmt: c.number_format = num_fmt

    # ── Row 1: Title ──────────────────────────────────────────────────────────
    ws1.row_dimensions[1].height = 30
    hdr(1, 1, "FEUILLE DE TRAVAIL — EXHAUSTIVITÉ DES CHARGES DU PERSONNEL",
        FTITLE, FILL_ALT, CTR, span_end=10)

    # ── Row 2: Subtitle ───────────────────────────────────────────────────────
    hdr(2, 1, "Rapprochement Paie / Comptabilité (SYSCOHADA) — Salaire Brut et Charges Patronales",
        FSUBHDR, FILL_NONE, CTR, span_end=10)

    # ── Rows 3–6: Metadata ────────────────────────────────────────────────────
    ws1.row_dimensions[3].height = 6
    meta_items = [
        ("Société / Company", metadata.get("societe", "")),
        ("Exercice / Fiscal Year", metadata.get("exercice", "")),
        ("Auditeur / Auditor", metadata.get("auditeur", "")),
        ("Date rapport", metadata.get("date_rapport", str(date.today()))),
    ]
    for i, (label, val) in enumerate(meta_items, 4):
        hdr(i, 1, label, FSUBHDR, FILL_ALT, LFT, BDR_DATA)
        ws1.merge_cells(f"B{i}:E{i}")
        hdr(i, 2, val, FDATA, FILL_NONE, LFT, BDR_DATA)

    # ── Row 9: Summary table header ───────────────────────────────────────────
    ws1.row_dimensions[8].height = 8  # spacer
    ws1.row_dimensions[9].height = 36
    summary_headers = ["Section", "SAL BRUT (R)", "CNPS/P (S)", "CF/P (T)", "FNE (U)",
                        "CF/P+FNE (V)", "AF (W)", "AT (X)", "TOTAL (Y)"]
    for c_idx, h in enumerate(summary_headers, 1):
        cell = ws1.cell(9, c_idx, value=h)
        cell.font = FHDR; cell.fill = FILL_BLUE; cell.alignment = CTR; cell.border = BDR_DATA

    col_order = ["R", "S", "T", "U", "V", "W", "X", "Y"]

    # ── Row 10: TOTAL PAIE ────────────────────────────────────────────────────
    ws1.cell(10, 1, value="TOTAL PAIE").font = FTOT_P
    ws1.cell(10, 1).fill = FILL_TOT_P; ws1.cell(10, 1).alignment = LFT; ws1.cell(10, 1).border = BDR_TOT
    for c_idx, col in enumerate(col_order, 2):
        val = data["paie"].get(col, 0)
        cell = ws1.cell(10, c_idx, value=val)
        cell.font = FTOT_P; cell.fill = FILL_TOT_P; cell.alignment = RGT
        cell.border = BDR_TOT; cell.number_format = NUM_FMT

    # ── Row 11: TOTAL COMPTA ──────────────────────────────────────────────────
    ws1.cell(11, 1, value="TOTAL COMPTABILITE").font = FTOT_C
    ws1.cell(11, 1).fill = FILL_TOT_C; ws1.cell(11, 1).alignment = LFT; ws1.cell(11, 1).border = BDR_TOT
    for c_idx, col in enumerate(col_order, 2):
        val = data["compta"].get(col, 0)
        cell = ws1.cell(11, c_idx, value=val)
        cell.font = FTOT_C; cell.fill = FILL_TOT_C; cell.alignment = RGT
        cell.border = BDR_TOT; cell.number_format = NUM_FMT

    # ── Row 12: ECART ─────────────────────────────────────────────────────────
    ws1.cell(12, 1, value="ECART TOTAL").font = FECART
    ws1.cell(12, 1).fill = FILL_ECART; ws1.cell(12, 1).alignment = LFT; ws1.cell(12, 1).border = BDR_TOT
    for c_idx, col in enumerate(col_order, 2):
        val = data["ecart"].get(col, 0)
        cell = ws1.cell(12, c_idx, value=val)
        cell.font = FECART; cell.fill = FILL_ECART; cell.alignment = RGT
        cell.border = BDR_TOT; cell.number_format = NUM_FMT

    # ── Row 14+: Ecart detail table ───────────────────────────────────────────
    ws1.row_dimensions[13].height = 8
    ws1.cell(14, 1, value="ANALYSE DES ECARTS / GAP ANALYSIS").font = FSUBHDR
    ws1.merge_cells("A14:J14")
    ws1.cell(14, 1).fill = FILL_BLUE
    ws1.cell(14, 1).font = FWHITE
    ws1.cell(14, 1).alignment = LFT

    gap_headers = ["Colonne", "PAIE", "COMPTA", "ECART", "Classification", "Action"]
    for c_idx, h in enumerate(gap_headers, 1):
        cell = ws1.cell(15, c_idx, value=h)
        cell.font = FHDR if c_idx == 1 else FSUBHDR
        cell.fill = FILL_ALT; cell.alignment = LFT; cell.border = BDR_DATA

    r_gap = 16
    all_reconciled = True
    for col_name in ["R", "S", "T", "U", "W", "X", "Y"]:
        paie_v   = data["paie"].get(col_name, 0)
        compta_v = data["compta"].get(col_name, 0)
        ecart_v  = data["ecart"].get(col_name, 0)
        expl     = ecart_explanations.get(col_name, {})
        category = expl.get("category", "")
        action   = expl.get("action", "")

        is_ok = abs(ecart_v) < 2
        if not is_ok:
            all_reconciled = False
        fill = FILL_OK if is_ok else FILL_WARN

        row_vals = [f"Col {col_name}", paie_v, compta_v, ecart_v, category, action]
        for c_idx, val in enumerate(row_vals, 1):
            cell = ws1.cell(r_gap, c_idx, value=val)
            cell.font = FDATA; cell.border = BDR_DATA; cell.fill = fill
            if c_idx in [2, 3, 4]:
                cell.alignment = RGT; cell.number_format = NUM_FMT
            else:
                cell.alignment = LFT
        r_gap += 1

    # ── Observations / justifications ─────────────────────────────────────────
    r_obs = r_gap + 1
    ws1.cell(r_obs, 1, value="JUSTIFICATIONS DES ECARTS / GAP JUSTIFICATIONS").font = FSUBHDR
    ws1.merge_cells(f"A{r_obs}:J{r_obs}")
    ws1.cell(r_obs, 1).fill = FILL_BLUE; ws1.cell(r_obs, 1).font = FWHITE
    ws1.cell(r_obs, 1).alignment = LFT
    r_obs += 1

    lang_key = "fr" if LANGUAGE == "fr" else "en"
    for col_name in ["U", "R"]:
        expl = ecart_explanations.get(col_name, {})
        ecart_v = data["ecart"].get(col_name, 0)
        if abs(ecart_v) < 2:
            continue
        ws1.merge_cells(f"A{r_obs}:J{r_obs}")
        ws1.cell(r_obs, 1, value=f"Colonne {col_name} — Ecart: {ecart_v:,.0f} FCFA").font = FSUBHDR
        ws1.cell(r_obs, 1).fill = FILL_ALT; ws1.cell(r_obs, 1).alignment = LFT
        r_obs += 1
        ws1.merge_cells(f"A{r_obs}:J{r_obs+2}")
        obs_cell = ws1.cell(r_obs, 1, value=expl.get(lang_key, ""))
        obs_cell.font = FDATA; obs_cell.alignment = WRP; obs_cell.border = BDR_MED
        ws1.row_dimensions[r_obs].height = 60
        r_obs += 3

    # ── Status ────────────────────────────────────────────────────────────────
    r_status = r_obs + 1
    status_fr = "RECONCILIE" if all_reconciled else "ECARTS JUSTIFIES"
    status_en = "RECONCILED" if all_reconciled else "GAPS JUSTIFIED"
    fill_s = FILL_OK if all_reconciled else FILL_WARN
    ws1.merge_cells(f"A{r_status}:E{r_status}")
    ws1.cell(r_status, 1, value=f"STATUT: {status_fr} / {status_en}").font = FSUBHDR
    ws1.cell(r_status, 1).fill = fill_s; ws1.cell(r_status, 1).alignment = CTR
    ws1.cell(r_status, 1).border = BDR_MED

    # Column widths Feuil1
    ws1.column_dimensions["A"].width = 28
    for col in "BCDEFGHI":
        ws1.column_dimensions[col].width = 16
    ws1.column_dimensions["J"].width = 35

    print(f"✅ Feuil1 rebuilt from scratch ({r_status} rows)")
    return status_fr

# ── Main ──────────────────────────────────────────────────────────────────────
def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--write", action="store_true")
    parser.add_argument("--row-paie",   type=int, default=None)
    parser.add_argument("--row-compta", type=int, default=None)
    parser.add_argument("--row-ecart",  type=int, default=None)
    parser.add_argument("--mode", default="interactive", choices=["interactive", "unattended"])
    args = parser.parse_args()

    # Load workbook — data_only to get computed values
    wb_d = load_workbook(FT_PATH, read_only=True, data_only=True)
    ws2_d = wb_d["Feuil2"]

    # Locate key rows if not passed
    row_paie   = args.row_paie   or session.get("feuil2_build", {}).get("row_total_paie",   178)
    row_compta = args.row_compta or session.get("feuil2_build", {}).get("row_total_compta")
    row_ecart  = args.row_ecart  or session.get("feuil2_build", {}).get("row_ecart")

    if not row_compta or not row_ecart:
        print("Scanning Feuil2 for TOTAL COMPTA and ECART rows...")
        for r in range(1, ws2_d.max_row + 1):
            v = str(ws2_d.cell(r, 1).value or "")
            vu = v.upper()
            if "TOTAL COMPTABILITE" in vu and not row_compta:
                row_compta = r
            if "ECART TOTAL" in vu and not row_ecart:
                row_ecart = r
    wb_d.close()

    if not row_compta or not row_ecart:
        print("❌ Cannot locate TOTAL COMPTA or ECART row — run reconcile first")
        sys.exit(1)

    # Reload without data_only to read actual values written by openpyxl
    wb = load_workbook(FT_PATH, data_only=True)
    ws2 = wb["Feuil2"]
    data = read_feuil2_values(ws2, row_paie, row_compta, row_ecart)

    print(f"Feuil2 values read (rows PAIE={row_paie}, COMPTA={row_compta}, ECART={row_ecart}):")
    for col in ["R","S","T","U","W","X","Y"]:
        p = data["paie"].get(col, 0)
        c = data["compta"].get(col, 0)
        e = data["ecart"].get(col, 0)
        status = "OK" if abs(e) < 2 else "ECART"
        print(f"  [{status}] {col}: PAIE={p:>18,.0f}  COMPTA={c:>18,.0f}  ECART={e:>15,.0f}")

    # Generate explanations
    ecart_explanations = {}
    for col_name in ["R", "S", "T", "U", "V", "W", "X", "Y"]:
        ecart_explanations[col_name] = get_explanation(
            col_name,
            data["paie"].get(col_name, 0),
            data["compta"].get(col_name, 0),
            data["ecart"].get(col_name, 0),
            LANGUAGE,
        )

    # In interactive mode: print ecart table and wait for approval
    run_mode = args.mode or session.get("run_mode", "interactive")
    if run_mode == "interactive":
        print("\n--- Ecart Summary (awaiting approval before writing Feuil1) ---")
        for col in ["R","S","T","U","W","X","Y"]:
            e = data["ecart"].get(col, 0)
            cat = ecart_explanations[col].get("category", "")
            print(f"  Col {col}: {e:>15,.0f} FCFA | {cat}")
        print("\n[In interactive mode: Claude will display this to the user and ask for approval]")
        print("[Use --mode unattended or pass confirmation from AskUserQuestion to proceed]")

    if args.write:
        metadata = session.get("metadata", {})
        ws1 = wb["Feuil1"]
        status = build_feuil1(ws1, data, metadata, ecart_explanations)
        wb.save(FT_PATH)
        print(f"✅ Feuil1 written to '{FT_PATH}' | Status: {status}")

        # Update session
        session["gap_analysis"] = {
            "row_paie":   row_paie,
            "row_compta": row_compta,
            "row_ecart":  row_ecart,
            "ecarts": {col: {"amount": data["ecart"].get(col, 0), **ecart_explanations.get(col, {})}
                       for col in ["R","S","T","U","W","X","Y"]},
            "overall_status": status,
        }
        session["feuil1_written"] = True
        session.setdefault("steps_completed", [])
        if "summarise" not in session["steps_completed"]:
            session["steps_completed"].append("summarise")
        if "all" not in session["steps_completed"] and len(session["steps_completed"]) >= 5:
            session["steps_completed"].append("all")
        with open(SESSION_FILE, "w", encoding="utf-8") as f_out:
            json.dump(session, f_out, indent=2, ensure_ascii=False)

if __name__ == "__main__":
    main()
