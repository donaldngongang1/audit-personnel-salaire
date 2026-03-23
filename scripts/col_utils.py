"""
col_utils.py — Dynamic column mapping for Feuil2.

Reads the Feuil2 header row and returns a dict mapping logical names to column indices.
This means formula generation and verification are driven by column *names*, not
hardcoded column letters — so if columns move tomorrow, nothing breaks.

Usage:
    from col_utils import get_feuil2_col_map, cols_to_sum_for_total, cols_to_sum_for_cf_fne

    col_map = get_feuil2_col_map(ws)
    # col_map["SAL_BRUT"] → 18  (or wherever "SAL BRUT" header is)
    # col_map["FNE"]      → 21
    # etc.
"""
from openpyxl.utils import get_column_letter as gcl

# ── Canonical logical names and the keywords that identify their header ────────
# Format: logical_name → list of keywords (ALL must appear in header for a match,
#         or first keyword suffices if it's unique enough)
# Matching is case-insensitive on the stripped header value.

COLUMN_MATCHERS = [
    # Logical name           Keywords to detect
    ("SAL_BRUT",            ["SAL BRUT"]),
    ("CNPS_P",              ["CNPS"]),          # catches "CNPS/P", "CNPS (Pension)"
    ("CF_P",                ["CF/P"]),           # but NOT "CF/P+FNE" — checked after V
    ("FNE",                 ["FNE"]),            # exact or as part of "CF/P+FNE" — see below
    ("CF_FNE",              ["CF/P+FNE"]),       # must be checked BEFORE CF_P to avoid stealing it
    ("AF",                  ["AF", "ALLOCATION FAMILIALE"]),
    ("AT",                  ["AT", "ACCIDENT DE TRAVAIL"]),
    ("TOTAL_COL",           ["TOTAL"]),          # Y column — the row-level total
]

# Order matters: CF_FNE must be checked before CF_P and FNE
_ORDERED_MATCHERS = [
    ("CF_FNE",    ["CF/P+FNE"]),
    ("SAL_BRUT",  ["SAL BRUT"]),
    ("CNPS_P",    ["CNPS/P", "CNPS"]),
    ("CF_P",      ["CF/P"]),            # won't match CF/P+FNE because that's already claimed
    ("FNE",       ["FNE"]),             # lone "FNE" header, not "CF/P+FNE"
    ("AF",        ["ALLOCATION FAMILIALE", " AF"]),  # space prefix to avoid matching "CAF"
    ("AT",        ["ACCIDENT DE TRAVAIL", " AT"]),
    ("TOTAL_COL", ["TOTAL"]),
]

# Blank/excluded columns (should never be written to)
BLANK_COL_KEYWORDS = ["SALAIRE BASE", "ANCIENNETE", "ANCIENNETÉ", "H. SUP", "H.SUP", "AUTRE GAIN"]


def get_feuil2_col_map(ws, header_row=3):
    """
    Scan `header_row` in `ws` and return a dict:
      { logical_name: col_index (1-based), ... }

    Also returns "BLANK_COLS": list of col indices that must stay blank.

    Raises ValueError if any required column is not found.
    """
    # Build raw map: stripped_upper_header → col_index
    raw = {}
    blank_cols = []
    for col in range(1, ws.max_column + 1):
        raw_val = ws.cell(header_row, col).value
        if raw_val is None:
            continue
        h = str(raw_val).strip().upper()
        raw[h] = col
        if any(kw in h for kw in ["SALAIRE BASE", "ANCIENNETE", "ANCIENNETÉ",
                                    "H. SUP", "H.SUP", "AUTRE GAIN"]):
            blank_cols.append(col)

    col_map = {"BLANK_COLS": blank_cols}

    # For each logical name, find the first matching header
    already_claimed = {}  # col_index → logical_name (prevent double assignment)

    for logical, keywords in _ORDERED_MATCHERS:
        for h, col_idx in raw.items():
            if col_idx in already_claimed:
                continue
            # Match: at least one keyword appears in the header
            if any(kw.upper() in h for kw in keywords):
                # Extra guard: for CF_P, reject if it contains "FNE" (that's CF_FNE)
                if logical == "CF_P" and "FNE" in h:
                    continue
                # Extra guard: for FNE (lone), reject if it also contains "CF"
                if logical == "FNE" and "CF" in h:
                    continue
                col_map[logical] = col_idx
                already_claimed[col_idx] = logical
                break

    # Validate all required columns were found
    required = ["SAL_BRUT", "CNPS_P", "CF_P", "FNE", "CF_FNE", "AF", "AT", "TOTAL_COL"]
    missing = [r for r in required if r not in col_map]
    if missing:
        found_headers = {h: c for h, c in raw.items() if c in already_claimed.values()}
        raise ValueError(
            f"Could not find columns for: {missing}\n"
            f"  Header row {header_row} contained: {dict(raw)}\n"
            f"  Matched so far: {col_map}"
        )

    return col_map


def cols_for_cf_fne(col_map):
    """Return (CF_P_col_idx, FNE_col_idx) — the two inputs for the V formula."""
    return col_map["CF_P"], col_map["FNE"]


def cols_for_row_total(col_map):
    """Return the list of col indices that sum to produce the row total (Y column).
    Excludes V (CF_FNE) because it is itself a sum of CF_P + FNE — adding it
    separately would double-count. The total is: SAL_BRUT + CNPS_P + CF_FNE + AF + AT.
    """
    return [
        col_map["SAL_BRUT"],
        col_map["CNPS_P"],
        col_map["CF_FNE"],   # V = CF/P + FNE combined
        col_map["AF"],
        col_map["AT"],
    ]


def all_numeric_cols(col_map):
    """Return all numeric col indices used in reconciliation (R through Y equiv)."""
    keys = ["SAL_BRUT", "CNPS_P", "CF_P", "FNE", "CF_FNE", "AF", "AT", "TOTAL_COL"]
    return [col_map[k] for k in keys if k in col_map]


def build_v_formula(col_map, row):
    """Build the V (CF/P+FNE) formula for a given row: =CF_P{r}+FNE{r}"""
    cf_col  = gcl(col_map["CF_P"])
    fne_col = gcl(col_map["FNE"])
    return f"={cf_col}{row}+{fne_col}{row}"


def build_y_formula(col_map, row):
    """Build the Y (TOTAL) formula for a given row.
    Uses CF_FNE (V) to avoid double-counting CF_P and FNE separately.
    Formula: =SAL_BRUT+CNPS_P+CF_FNE+AF+AT
    """
    col_indices = cols_for_row_total(col_map)
    terms = "+".join(f"{gcl(c)}{row}" for c in col_indices)
    return f"={terms}"


def build_sum_formula(col_map, logical_name, start_row, end_row):
    """Build a SUM formula for one logical column: =SUM(col{start}:col{end})"""
    col_idx = col_map[logical_name]
    col_letter = gcl(col_idx)
    return f"=SUM({col_letter}{start_row}:{col_letter}{end_row})"


def build_total_compta_formula(col_map, logical_name, subtot_rows):
    """Build TOTAL COMPTA formula for one column: =col{stA}+col{stB}+col{stC}"""
    col_letter = gcl(col_map[logical_name])
    terms = "+".join(f"{col_letter}{r}" for r in subtot_rows)
    return f"={terms}"


def build_ecart_formula(col_map, logical_name, compta_row, paie_row):
    """Build ECART formula: =col{compta}-col{paie}"""
    col_letter = gcl(col_map[logical_name])
    return f"={col_letter}{compta_row}-{col_letter}{paie_row}"


def describe_col_map(col_map):
    """Pretty-print the resolved column map for debugging."""
    lines = ["Feuil2 Column Map (resolved from header row):"]
    for name, val in col_map.items():
        if name == "BLANK_COLS":
            lines.append(f"  BLANK_COLS: {[gcl(c) for c in val]}")
        else:
            lines.append(f"  {name:15s} → col {val:2d} ({gcl(val)})")
    return "\n".join(lines)
